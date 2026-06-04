"""
generate_sponsor_files.py
Synthesizes all TCSW sponsor data from research sources into:
  1. TCSW_Sponsors_Master.xlsx  — multi-sheet Excel workbook
  2. TCSW_Sponsors_Master.pdf   — nicely formatted PDF report
"""

import os
import sys
import textwrap
from datetime import datetime

import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer,
    HRFlowable, PageBreak, KeepTogether
)
from reportlab.platypus.flowables import HRFlowable
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

# ──────────────────────────────────────────────────────────────────────────────
# MASTER DATA  (synthesised from all research sources)
# ──────────────────────────────────────────────────────────────────────────────

# Colour palette (TCSW brand-ish)
TCSW_TEAL   = "005F73"
TCSW_GOLD   = "EE9B00"
TCSW_LIGHT  = "E9F5F7"
TCSW_MID    = "94D2BD"
WHITE       = "FFFFFF"
DARK_TEXT   = "001219"

SPONSORS = [
    # ── 2026 confirmed ──────────────────────────────────────────────────────
    {
        "year": 2026, "organization": "BETA.MN (The Beta Group)",
        "tier_role": "Organizer / Operator", "category": "Nonprofit",
        "sector": "Nonprofit / Community",
        "contact": "Angela Eifert — Executive Director",
        "email": "Angela.Eifert@BETA.MN", "website": "https://www.beta.mn",
        "named_event": "", "venue_sponsor": False, "multi_year": True,
        "years_confirmed": "2014–2026",
        "notes": "501(c)(3) nonprofit; primary operator of TCSW every year",
        "sources": "tcstartupweek.com; beta.mn",
    },
    {
        "year": 2026, "organization": "Communiful",
        "tier_role": "Co-Host / Co-Organizer", "category": "Tech / Community",
        "sector": "Coworking / Creative Hub",
        "contact": "Hayley Brooks — Founder",
        "email": "", "website": "https://communiful.com",
        "named_event": "", "venue_sponsor": False, "multi_year": True,
        "years_confirmed": "2025–2026",
        "notes": "Creative hub & coworking community; co-hosting 2025 & 2026",
        "sources": "tcstartupweek.com; communiful.com",
    },
    {
        "year": 2026, "organization": "JPMorgan Chase",
        "tier_role": "Premier Sponsor", "category": "Finance / Banking",
        "sector": "Financial Services",
        "contact": "", "email": "", "website": "https://jpmorganchase.com",
        "named_event": "", "venue_sponsor": False, "multi_year": True,
        "years_confirmed": "2023–2026",
        "notes": "BETA Showcase sponsor 2023; listed 2026 sponsor",
        "sources": "organizations_database.json; SPONSORS_DATABASE.md",
    },
    {
        "year": 2026, "organization": "Groove Capital",
        "tier_role": "Event Partner — Pitch Night", "category": "VC / Investor",
        "sector": "Venture Capital",
        "contact": "", "email": "", "website": "https://groovecapital.com",
        "named_event": "Pitch Night (portfolio showcase)", "venue_sponsor": False, "multi_year": True,
        "years_confirmed": "2025–2026",
        "notes": "MN-based VC; hosts portfolio showcase pitch night",
        "sources": "SPONSORS_DATABASE.csv; organizations_database.json",
    },
    {
        "year": 2026, "organization": "Bread & Butter Ventures",
        "tier_role": "Ecosystem Partner / VC Panel", "category": "VC / Investor",
        "sector": "Venture Capital",
        "contact": "", "email": "", "website": "https://breadandbutter.vc",
        "named_event": "VC Panel", "venue_sponsor": False, "multi_year": True,
        "years_confirmed": "2025–2026",
        "notes": "MN VC; participated in VC panel; listed 2026 sponsor",
        "sources": "SPONSORS_DATABASE.csv; organizations_database.json",
    },
    {
        "year": 2026, "organization": "Augeo",
        "tier_role": "Sponsor", "category": "Technology / Loyalty",
        "sector": "Technology",
        "contact": "", "email": "", "website": "https://augeo.com",
        "named_event": "", "venue_sponsor": False, "multi_year": False,
        "years_confirmed": "2026",
        "notes": "Listed 2026 sponsor",
        "sources": "organizations_database.json; SPONSORS_DATABASE.csv",
    },
    {
        "year": 2026, "organization": "Northeast Bank",
        "tier_role": "Sponsor", "category": "Banking / Finance",
        "sector": "Financial Services",
        "contact": "", "email": "", "website": "https://northeastbank.com",
        "named_event": "", "venue_sponsor": False, "multi_year": False,
        "years_confirmed": "2026",
        "notes": "Listed 2026 sponsor",
        "sources": "organizations_database.json",
    },
    {
        "year": 2026, "organization": "Taft Law",
        "tier_role": "Sponsor — Legal", "category": "Legal / Law Firm",
        "sector": "Professional Services",
        "contact": "", "email": "", "website": "https://taftlaw.com",
        "named_event": "Exit Planning Sessions", "venue_sponsor": False, "multi_year": False,
        "years_confirmed": "2026",
        "notes": "Annual exit planning legal sessions",
        "sources": "organizations_database.json; SPONSORS_DATABASE.csv",
    },
    {
        "year": 2026, "organization": "Mairs & Power",
        "tier_role": "Sponsor / Panel", "category": "Investment Management",
        "sector": "Financial Services",
        "contact": "Scott Burns — Managing Director (Venture Capital division)",
        "email": "", "website": "https://mairsandpower.com",
        "named_event": "B2B Scaling Panel (Thursday)", "venue_sponsor": False, "multi_year": True,
        "years_confirmed": "2025–2026",
        "notes": "Scott Burns spoke on Thursday B2B scaling panel in 2025",
        "sources": "SPONSORS_DATABASE.csv; organizations_database.json",
    },
    {
        "year": 2026, "organization": "Brooksource",
        "tier_role": "Sponsor", "category": "Staffing / Technology",
        "sector": "Professional Services",
        "contact": "", "email": "", "website": "https://brooksource.com",
        "named_event": "", "venue_sponsor": False, "multi_year": False,
        "years_confirmed": "2026",
        "notes": "Listed 2026 sponsor",
        "sources": "organizations_database.json",
    },
    {
        "year": 2026, "organization": "Shakopee Mdewakanton Sioux Community",
        "tier_role": "Sponsor", "category": "Tribal Nation / Community",
        "sector": "Government / Tribal",
        "contact": "", "email": "", "website": "https://shakopeedakota.org",
        "named_event": "", "venue_sponsor": False, "multi_year": False,
        "years_confirmed": "2026",
        "notes": "Tribal nation sponsor — MN community partner",
        "sources": "organizations_database.json",
    },
    {
        "year": 2026, "organization": "Idea Fund",
        "tier_role": "Sponsor / Happy Hour", "category": "Regional VC (La Crosse, WI)",
        "sector": "Venture Capital",
        "contact": "", "email": "", "website": "",
        "named_event": "Monday Happy Hour", "venue_sponsor": False, "multi_year": True,
        "years_confirmed": "2025–2026",
        "notes": "Regional VC; hosted Monday happy hour 2025",
        "sources": "SPONSORS_DATABASE.csv; organizations_database.json",
    },
    {
        "year": 2026, "organization": "Biometrica Health",
        "tier_role": "Sponsor", "category": "Healthcare Technology",
        "sector": "Health Tech",
        "contact": "", "email": "", "website": "",
        "named_event": "", "venue_sponsor": False, "multi_year": False,
        "years_confirmed": "2026",
        "notes": "Listed 2026 sponsor",
        "sources": "organizations_database.json",
    },
    {
        "year": 2026, "organization": "ezer21",
        "tier_role": "Sponsor", "category": "Technology / Startup",
        "sector": "Technology",
        "contact": "", "email": "", "website": "",
        "named_event": "", "venue_sponsor": False, "multi_year": False,
        "years_confirmed": "2026",
        "notes": "Listed 2026 sponsor",
        "sources": "organizations_database.json",
    },
    {
        "year": 2026, "organization": "Full Stack Saint Paul Scale Up",
        "tier_role": "Sponsor / Program", "category": "Startup Support",
        "sector": "Ecosystem / Program",
        "contact": "", "email": "", "website": "",
        "named_event": "", "venue_sponsor": False, "multi_year": False,
        "years_confirmed": "2026",
        "notes": "Scale-up program; listed 2026 sponsor",
        "sources": "organizations_database.json",
    },
    {
        "year": 2026, "organization": "MN Cup / University of Minnesota",
        "tier_role": "Competition Partner", "category": "Competition / University",
        "sector": "Education / Competition",
        "contact": "", "email": "hellomncup@umn.edu", "website": "https://mncup.org",
        "named_event": "Grand Finale ($400K+ prizes)", "venue_sponsor": False, "multi_year": True,
        "years_confirmed": "2014–2026",
        "notes": "Annual pitch competition Grand Finale held during TCSW every year",
        "sources": "SPONSORS_DATABASE.csv; ALL_YEARS_SPONSOR_SUMMARY.md",
    },
    {
        "year": 2026, "organization": "Launch Minnesota",
        "tier_role": "Government Partner", "category": "Government / Economic Dev",
        "sector": "Government",
        "contact": "", "email": "", "website": "https://launchmn.com",
        "named_event": "Tuesday Happy Hour", "venue_sponsor": False, "multi_year": True,
        "years_confirmed": "2019–2026",
        "notes": "State of MN startup initiative; hosts happy hour annually",
        "sources": "SPONSORS_DATABASE.csv; organizations_database.json",
    },
    # ── 2025 only ────────────────────────────────────────────────────────────
    {
        "year": 2025, "organization": "Great North Ventures",
        "tier_role": "Event Partner — Investor Networking", "category": "VC / Investor",
        "sector": "Venture Capital",
        "contact": "", "email": "", "website": "https://greatnorthventures.com",
        "named_event": "Investor-Founder Networking Event (Tuesday)", "venue_sponsor": False, "multi_year": False,
        "years_confirmed": "2025",
        "notes": "MN VC; co-hosted Tuesday investor-founder networking",
        "sources": "SPONSORS_DATABASE.csv; start-midwest.com",
    },
    {
        "year": 2025, "organization": "EmpowerHER",
        "tier_role": "Community Partner — Investor Networking", "category": "Organization",
        "sector": "Community / DEI",
        "contact": "", "email": "", "website": "",
        "named_event": "Investor-Founder Networking Event (Tuesday)", "venue_sponsor": False, "multi_year": False,
        "years_confirmed": "2025",
        "notes": "Women founders program; co-hosted Tuesday networking",
        "sources": "SPONSORS_DATABASE.csv",
    },
    {
        "year": 2025, "organization": "gener8tor",
        "tier_role": "Event Partner — Showcase", "category": "Accelerator",
        "sector": "Startup Accelerator",
        "contact": "", "email": "", "website": "https://gener8tor.com",
        "named_event": "BETA × gener8tor Showcase (200+ attendees)", "venue_sponsor": False, "multi_year": False,
        "years_confirmed": "2025",
        "notes": "Startup accelerator; co-hosted showcase with 200+ attendees",
        "sources": "SPONSORS_DATABASE.csv; twincitiescollective.com",
    },
    {
        "year": 2025, "organization": "Matchstick Ventures",
        "tier_role": "VC Panel Participant", "category": "VC / Investor",
        "sector": "Venture Capital",
        "contact": "", "email": "", "website": "https://matchstickventures.com",
        "named_event": "Monday VC Panel", "venue_sponsor": False, "multi_year": False,
        "years_confirmed": "2025",
        "notes": "MN VC; participated in Monday VC panel",
        "sources": "SPONSORS_DATABASE.csv; start-midwest.com",
    },
    {
        "year": 2025, "organization": "M25",
        "tier_role": "VC Panel Participant", "category": "VC / Investor (Chicago)",
        "sector": "Venture Capital",
        "contact": "", "email": "", "website": "",
        "named_event": "Monday VC Panel", "venue_sponsor": False, "multi_year": False,
        "years_confirmed": "2025",
        "notes": "Chicago-based VC; participated in Monday VC panel",
        "sources": "SPONSORS_DATABASE.csv; start-midwest.com",
    },
    {
        "year": 2025, "organization": "Hyde Park Venture Partners",
        "tier_role": "VC Panel Participant", "category": "VC / Investor",
        "sector": "Venture Capital",
        "contact": "", "email": "", "website": "",
        "named_event": "Monday VC Panel", "venue_sponsor": False, "multi_year": False,
        "years_confirmed": "2025",
        "notes": "Participated in Monday VC panel",
        "sources": "SPONSORS_DATABASE.csv; start-midwest.com",
    },
    {
        "year": 2025, "organization": "DEED (MN Dept of Employment & Economic Dev)",
        "tier_role": "Government Partner — Resource Fair", "category": "Government",
        "sector": "Government",
        "contact": "", "email": "", "website": "",
        "named_event": "Thursday St. Paul Resource Fair", "venue_sponsor": False, "multi_year": True,
        "years_confirmed": "2019, 2025",
        "notes": "State agency providing startup resources",
        "sources": "SPONSORS_DATABASE.csv; twincitiescollective.com",
    },
    {
        "year": 2025, "organization": "Naturally MN",
        "tier_role": "Event Partner — Food & Bev Showcase", "category": "Industry Association",
        "sector": "Food & Beverage",
        "contact": "", "email": "", "website": "",
        "named_event": "Innovation Crawl co-event + F&B Showcase (Tuesday)", "venue_sponsor": False, "multi_year": False,
        "years_confirmed": "2025",
        "notes": "Food & beverage industry showcase during Innovation Crawl",
        "sources": "SPONSORS_DATABASE.csv; start-midwest.com",
    },
    {
        "year": 2025, "organization": "MN:EVC",
        "tier_role": "Ecosystem Partner — Networking", "category": "Investor Network",
        "sector": "Venture Capital",
        "contact": "", "email": "", "website": "",
        "named_event": "Networking Event", "venue_sponsor": False, "multi_year": False,
        "years_confirmed": "2025",
        "notes": "Minnesota Early-stage Venture Capital network",
        "sources": "SPONSORS_DATABASE.md",
    },
    # ── 2023 ─────────────────────────────────────────────────────────────────
    {
        "year": 2023, "organization": "Optum (UnitedHealth Group)",
        "tier_role": "Title Sponsor — Cup Naming Rights", "category": "Healthcare Technology",
        "sector": "Health Tech",
        "contact": "", "email": "", "website": "https://optum.com",
        "named_event": "Optum Cup (healthcare pitch competition)", "venue_sponsor": False, "multi_year": False,
        "years_confirmed": "2023",
        "notes": "Highest-level sponsor 2023; named the healthcare pitch competition 'Optum Cup'",
        "sources": "SPONSORS_DATABASE.csv; tcsw2023.sched.com",
    },
    {
        "year": 2023, "organization": "TriNet",
        "tier_role": "Presenting Sponsor — VC Fireside Chat", "category": "HR Technology",
        "sector": "HR / Benefits",
        "contact": "", "email": "", "website": "https://trinet.com",
        "named_event": "VC Fireside Chat", "venue_sponsor": False, "multi_year": False,
        "years_confirmed": "2023",
        "notes": "HR/benefits platform for startups; hosted VC Fireside Chat; branding on stage",
        "sources": "SPONSORS_DATABASE.csv; tcsw2023.sched.com",
    },
    {
        "year": 2023, "organization": "Twin Ignition",
        "tier_role": "Presenting Sponsor — Opening Party", "category": "Startup Support Org",
        "sector": "Ecosystem / Community",
        "contact": "", "email": "", "website": "https://twinignition.com",
        "named_event": "Opening Party (Keg House + Startup Garage)", "venue_sponsor": True, "multi_year": True,
        "years_confirmed": "2022–2024",
        "notes": "Co-hosted opening party; provided Startup Garage + Keg House venues",
        "sources": "SPONSORS_DATABASE.csv; organizations_database.json",
    },
    {
        "year": 2023, "organization": "Workbox",
        "tier_role": "Venue Sponsor", "category": "Coworking",
        "sector": "Coworking / Real Estate",
        "contact": "", "email": "", "website": "https://workbox.com",
        "named_event": "Multiple session venues", "venue_sponsor": True, "multi_year": True,
        "years_confirmed": "2022–2024",
        "notes": "Premium coworking space; hosted multiple TCSW sessions",
        "sources": "SPONSORS_DATABASE.csv; tcsw2023.sched.com",
    },
    {
        "year": 2023, "organization": "Augurian & Cloudburst SBC",
        "tier_role": "Venue Sponsor", "category": "Marketing / Coworking",
        "sector": "Marketing / Agency",
        "contact": "", "email": "", "website": "https://augurian.com",
        "named_event": "Multiple session venues", "venue_sponsor": True, "multi_year": False,
        "years_confirmed": "2023",
        "notes": "Marketing agency + co-branded coworking space",
        "sources": "SPONSORS_DATABASE.csv; tcsw2023.sched.com",
    },
    {
        "year": 2023, "organization": "FINNOVATION Lab",
        "tier_role": "Venue Sponsor", "category": "Fintech Innovation",
        "sector": "Fintech",
        "contact": "", "email": "", "website": "https://finnovationlab.com",
        "named_event": "Multiple fintech sessions", "venue_sponsor": True, "multi_year": True,
        "years_confirmed": "2022–2024",
        "notes": "Fintech innovation lab; hosted fintech track sessions",
        "sources": "SPONSORS_DATABASE.csv; tcsw2023.sched.com",
    },
    {
        "year": 2023, "organization": "Lab651 HQ",
        "tier_role": "Venue Sponsor", "category": "Tech Innovation",
        "sector": "Technology",
        "contact": "", "email": "", "website": "https://lab651.com",
        "named_event": "AI + software sessions", "venue_sponsor": True, "multi_year": False,
        "years_confirmed": "2023",
        "notes": "Tech innovation lab; hosted AI + software sessions",
        "sources": "SPONSORS_DATABASE.csv; tcsw2023.sched.com",
    },
    {
        "year": 2023, "organization": "Luminary Arts Center",
        "tier_role": "Venue Partner — Main Stage", "category": "Arts / Event Space",
        "sector": "Venue",
        "contact": "", "email": "", "website": "https://luminaryartscenter.com",
        "named_event": "Main stage venue (Community Stage + Bemis Room)", "venue_sponsor": True, "multi_year": False,
        "years_confirmed": "2023",
        "notes": "Primary main stage venue for 2023 TCSW",
        "sources": "SPONSORS_DATABASE.csv; tcsw2023.sched.com",
    },
    {
        "year": 2023, "organization": "West Monroe",
        "tier_role": "Venue / Session Sponsor", "category": "Consulting / Technology",
        "sector": "Consulting",
        "contact": "", "email": "", "website": "https://westmonroe.com",
        "named_event": "AI Day session host", "venue_sponsor": True, "multi_year": False,
        "years_confirmed": "2023",
        "notes": "Digital consulting firm; hosted AI Day sessions",
        "sources": "organizations_database.json",
    },
    {
        "year": 2023, "organization": "Stratasys",
        "tier_role": "Venue / Session Sponsor", "category": "3D Printing / Manufacturing",
        "sector": "Manufacturing / Tech",
        "contact": "", "email": "", "website": "https://stratasys.com",
        "named_event": "3D Printing session host", "venue_sponsor": True, "multi_year": False,
        "years_confirmed": "2023",
        "notes": "3D printing company; hosted manufacturing/3D printing sessions",
        "sources": "organizations_database.json",
    },
    # ── 2019 (Fortune 500 peak year) ─────────────────────────────────────────
    {
        "year": 2019, "organization": "Target Corporation",
        "tier_role": "Title Sponsor", "category": "Retail / Corporate (Fortune 500)",
        "sector": "Retail",
        "contact": "Gene Han — VP New Ventures & Accelerators",
        "email": "", "website": "https://target.com",
        "named_event": "Kickoff Event + Retail Track (Target Plaza Commons)", "venue_sponsor": False, "multi_year": False,
        "years_confirmed": "2019",
        "notes": "First Fortune 500 title sponsor year; hosted kickoff + retail track at HQ",
        "sources": "SPONSORS_DATABASE.csv; medium.com/twin-cities-startup-week; FORTUNE500_SPONSORS_2019.md",
    },
    {
        "year": 2019, "organization": "Cargill",
        "tier_role": "Title Sponsor", "category": "Agriculture / Corporate (Fortune 500)",
        "sector": "Agriculture / Food",
        "contact": "Lawrence Wang — Digitalization & Analytics Strategy Lead",
        "email": "", "website": "https://cargill.com",
        "named_event": "Techstars Farm to Fork Demo Day (Oct 15)", "venue_sponsor": False, "multi_year": True,
        "years_confirmed": "2018–2019",
        "notes": "Wayzata MN HQ; co-sponsored Farm to Fork; title sponsor 2019",
        "sources": "SPONSORS_DATABASE.csv; medium.com/twin-cities-startup-week; FORTUNE500_SPONSORS_2019.md",
    },
    {
        "year": 2019, "organization": "3M",
        "tier_role": "Title Sponsor", "category": "Manufacturing / Corporate (Fortune 500)",
        "sector": "Manufacturing / Science",
        "contact": "Ben Wright — Director of 3M Ventures",
        "email": "", "website": "https://3m.com",
        "named_event": "BETA Backers Program", "venue_sponsor": False, "multi_year": True,
        "years_confirmed": "2018–2019",
        "notes": "Maplewood MN HQ; participated in BETA Backers (investors meet early-stage startups)",
        "sources": "SPONSORS_DATABASE.csv; medium.com/twin-cities-startup-week; FORTUNE500_SPONSORS_2019.md",
    },
    {
        "year": 2019, "organization": "Microsoft",
        "tier_role": "Corporate Sponsor", "category": "Technology",
        "sector": "Technology",
        "contact": "", "email": "", "website": "https://microsoft.com",
        "named_event": "", "venue_sponsor": False, "multi_year": False,
        "years_confirmed": "2019",
        "notes": "Confirmed in Twin Cities Business article",
        "sources": "SPONSORS_DATABASE.csv; tcbmag.com",
    },
    {
        "year": 2019, "organization": "AWS (Amazon Web Services)",
        "tier_role": "Corporate Sponsor", "category": "Technology / Cloud",
        "sector": "Technology",
        "contact": "", "email": "", "website": "https://aws.amazon.com",
        "named_event": "", "venue_sponsor": False, "multi_year": False,
        "years_confirmed": "2019",
        "notes": "Confirmed in Twin Cities Business article; AWS Activate program alignment",
        "sources": "SPONSORS_DATABASE.csv; tcbmag.com",
    },
    {
        "year": 2019, "organization": "Ecolab",
        "tier_role": "Event Co-Sponsor", "category": "Industrial / Environmental",
        "sector": "Industrial / Sustainability",
        "contact": "", "email": "", "website": "",
        "named_event": "Farm to Fork co-sponsor (with Cargill)", "venue_sponsor": False, "multi_year": False,
        "years_confirmed": "2019",
        "notes": "Co-sponsored Techstars Farm to Fork with Cargill",
        "sources": "SPONSORS_DATABASE.csv; tcbmag.com",
    },
    {
        "year": 2019, "organization": "GreaterMSP",
        "tier_role": "Presenting / Promotional Partner", "category": "Economic Development",
        "sector": "Economic Development",
        "contact": "", "email": "", "website": "",
        "named_event": "", "venue_sponsor": False, "multi_year": False,
        "years_confirmed": "2019",
        "notes": "Regional economic development organization; promotional partner",
        "sources": "SPONSORS_DATABASE.csv; greatermsp.org",
    },
    {
        "year": 2019, "organization": "Make It. MSP",
        "tier_role": "Fly-In Program Sponsor", "category": "Economic Development / Talent",
        "sector": "Economic Development",
        "contact": "", "email": "", "website": "",
        "named_event": "Fly-In Program (non-MN job seekers)", "venue_sponsor": False, "multi_year": True,
        "years_confirmed": "2018–2019",
        "notes": "Flew in non-Minnesota job seekers to experience TCSW",
        "sources": "SPONSORS_DATABASE.csv; tcbmag.com",
    },
    {
        "year": 2019, "organization": "Branch Messenger",
        "tier_role": "Fly-In Program Sponsor", "category": "Tech Startup (now part of Square)",
        "sector": "Technology",
        "contact": "", "email": "", "website": "",
        "named_event": "Fly-In Program (non-MN job seekers)", "venue_sponsor": False, "multi_year": True,
        "years_confirmed": "2018–2019",
        "notes": "Minneapolis-based startup; later acquired by Square/Block",
        "sources": "SPONSORS_DATABASE.csv; tcbmag.com",
    },
    {
        "year": 2019, "organization": "Techstars",
        "tier_role": "Event Partner", "category": "Accelerator",
        "sector": "Startup Accelerator",
        "contact": "", "email": "", "website": "",
        "named_event": "Farm to Fork Demo Day", "venue_sponsor": False, "multi_year": False,
        "years_confirmed": "2019",
        "notes": "Techstars Farm to Fork accelerator Demo Day hosted during TCSW",
        "sources": "SPONSORS_DATABASE.csv; tcbmag.com",
    },
    # ── 2018 ─────────────────────────────────────────────────────────────────
    {
        "year": 2018, "organization": "Mayo Clinic",
        "tier_role": "MANOVA Summit Sponsor", "category": "Healthcare",
        "sector": "Healthcare",
        "contact": "", "email": "", "website": "",
        "named_event": "MANOVA Summit (health innovation)", "venue_sponsor": False, "multi_year": False,
        "years_confirmed": "2018",
        "notes": "Backed health innovation summit co-located with TCSW",
        "sources": "SPONSORS_DATABASE.csv; tcbmag.com",
    },
    {
        "year": 2018, "organization": "Walmart",
        "tier_role": "MANOVA Summit Sponsor", "category": "Retail / Corporate",
        "sector": "Retail",
        "contact": "", "email": "", "website": "",
        "named_event": "MANOVA Summit (health innovation)", "venue_sponsor": False, "multi_year": False,
        "years_confirmed": "2018",
        "notes": "Backed health innovation summit co-located with TCSW",
        "sources": "SPONSORS_DATABASE.csv; tcbmag.com",
    },
    {
        "year": 2018, "organization": "Lyft",
        "tier_role": "Transport Sponsor", "category": "Transportation / Tech",
        "sector": "Technology",
        "contact": "", "email": "", "website": "https://lyft.com",
        "named_event": "Free Lyft shuttle service during TCSW week", "venue_sponsor": False, "multi_year": False,
        "years_confirmed": "2018",
        "notes": "Provided free Lyft bus shuttle for attendees during entire week",
        "sources": "organizations_database.json",
    },
    # ── 2017 ─────────────────────────────────────────────────────────────────
    {
        "year": 2017, "organization": "Allianz Life",
        "tier_role": "Major Sponsor — Innovation Day Host", "category": "Insurance / Finance",
        "sector": "Insurance / Financial Services",
        "contact": "", "email": "", "website": "",
        "named_event": "Allianz Life Ventures Innovation Day 2017", "venue_sponsor": True, "multi_year": False,
        "years_confirmed": "2017",
        "notes": "Hosted Innovation Day at Allianz Life HQ; major corporate sponsor",
        "sources": "SPONSORS_DATABASE.csv; twincitiesstartupweek2017.sched.com",
    },
    {
        "year": 2017, "organization": "Fredrikson & Byron P.A.",
        "tier_role": "Legal Sponsor", "category": "Law Firm",
        "sector": "Professional Services",
        "contact": "", "email": "", "website": "https://fredlaw.com",
        "named_event": "Angel Investing 101 + Financing a Startup sessions", "venue_sponsor": False, "multi_year": True,
        "years_confirmed": "2017, 2023",
        "notes": "Perennial legal sponsor; hosted finance/legal sessions",
        "sources": "SPONSORS_DATABASE.csv; twincitiesstartupweek2017.sched.com",
    },
    {
        "year": 2017, "organization": "Avisen Legal P.A.",
        "tier_role": "Legal Sponsor", "category": "Law Firm",
        "sector": "Professional Services",
        "contact": "", "email": "", "website": "",
        "named_event": "Open Office Hours legal counsel", "venue_sponsor": False, "multi_year": False,
        "years_confirmed": "2017",
        "notes": "Provided free open office hours legal counsel for startups",
        "sources": "SPONSORS_DATABASE.csv; twincitiesstartupweek2017.sched.com",
    },
    {
        "year": 2017, "organization": "CoCo (Coworking Community)",
        "tier_role": "Venue Partner — Co-Working", "category": "Coworking",
        "sector": "Coworking",
        "contact": "", "email": "", "website": "",
        "named_event": "Free co-working (Downtown + St. Paul)", "venue_sponsor": True, "multi_year": False,
        "years_confirmed": "2017",
        "notes": "Provided free co-working space at downtown Minneapolis + St. Paul locations",
        "sources": "SPONSORS_DATABASE.csv",
    },
    {
        "year": 2017, "organization": "Industrious",
        "tier_role": "Venue Partner — Co-Working", "category": "Coworking",
        "sector": "Coworking",
        "contact": "", "email": "", "website": "",
        "named_event": "Free co-working space during TCSW week", "venue_sponsor": True, "multi_year": False,
        "years_confirmed": "2017",
        "notes": "Provided free co-working to attendees",
        "sources": "SPONSORS_DATABASE.csv",
    },
    {
        "year": 2017, "organization": "WeWork",
        "tier_role": "Venue Partner — Co-Working", "category": "Coworking",
        "sector": "Coworking",
        "contact": "", "email": "", "website": "https://wework.com",
        "named_event": "Free co-working + 2023 NorthLoop location", "venue_sponsor": True, "multi_year": True,
        "years_confirmed": "2017, 2023",
        "notes": "Free co-working 2017; NorthLoop Minneapolis location 2023",
        "sources": "SPONSORS_DATABASE.csv; organizations_database.json",
    },
    {
        "year": 2017, "organization": "Mall of America",
        "tier_role": "Venue Partner", "category": "Retail / Venue",
        "sector": "Retail / Events",
        "contact": "", "email": "", "website": "",
        "named_event": "Night at the Mall event", "venue_sponsor": True, "multi_year": False,
        "years_confirmed": "2017",
        "notes": "Hosted 'Night at the Mall' networking event",
        "sources": "SPONSORS_DATABASE.csv",
    },
    {
        "year": 2017, "organization": "Bauhaus Brewery",
        "tier_role": "Venue Partner", "category": "Food & Beverage / Venue",
        "sector": "Food & Beverage",
        "contact": "", "email": "", "website": "",
        "named_event": "Drone racing event", "venue_sponsor": True, "multi_year": True,
        "years_confirmed": "2017, 2023",
        "notes": "Hosted drone racing event 2017; Medical Device Networking happy hour 2023",
        "sources": "SPONSORS_DATABASE.csv",
    },
    # ── Founding / Perennial ─────────────────────────────────────────────────
    {
        "year": 2014, "organization": "TECHdotMN",
        "tier_role": "Founding Partner", "category": "Media / Ecosystem",
        "sector": "Media / Technology",
        "contact": "", "email": "", "website": "",
        "named_event": "", "venue_sponsor": False, "multi_year": True,
        "years_confirmed": "2014–present",
        "notes": "Founding community media partner; annual supporter",
        "sources": "ALL_YEARS_SPONSOR_SUMMARY.md",
    },
    {
        "year": 2014, "organization": "Fueled Collective",
        "tier_role": "Founding Partner", "category": "Coworking",
        "sector": "Coworking",
        "contact": "", "email": "", "website": "",
        "named_event": "", "venue_sponsor": True, "multi_year": False,
        "years_confirmed": "2014",
        "notes": "Founding co-working partner; Year 1 (2014)",
        "sources": "ALL_YEARS_SPONSOR_SUMMARY.md",
    },
    {
        "year": 2014, "organization": "Minnestar",
        "tier_role": "Founding Partner", "category": "Community Org",
        "sector": "Ecosystem / Community",
        "contact": "", "email": "", "website": "",
        "named_event": "", "venue_sponsor": False, "multi_year": False,
        "years_confirmed": "2014",
        "notes": "Founding community partner; organizers of Minnebar",
        "sources": "ALL_YEARS_SPONSOR_SUMMARY.md",
    },
]

# 2026 Sponsorship tiers
TIERS_2026 = [
    {"tier": "Premier",   "min_investment": "TBD",    "benefits": "Maximum stage presence + activation package + all lower tier benefits"},
    {"tier": "Presenter", "min_investment": "TBD",    "benefits": "Enhanced digital presence + dedicated session + activation + Builder benefits"},
    {"tier": "Host",      "min_investment": "TBD",    "benefits": "Hosted activation + session programming rights + digital presence + Builder benefits"},
    {"tier": "Builder",   "min_investment": "$5,000", "benefits": "Stage recognition, digital presence, logo on materials, website listing"},
]

# Key contacts
CONTACTS = [
    {"name": "Angela Eifert", "title": "Executive Director, BETA.MN", "email": "Angela.Eifert@BETA.MN", "note": "Primary sponsorship contact for TCSW 2026"},
    {"name": "info@twincitiesstartupweek.com", "title": "General Inquiry", "email": "info@twincitiesstartupweek.com", "note": "General TCSW contact"},
    {"name": "Hayley Brooks", "title": "Founder, Communiful", "email": "", "note": "Co-host 2025–2026"},
]

# 2026 event structure
EVENT_DAYS = [
    {"day": "Day 1 — Mon Sept 14", "focus": "Registration + MN Cup Grand Finale", "sponsor_opportunity": "Registration sponsor; MN Cup naming rights"},
    {"day": "Day 2 — Tue Sept 15", "focus": "Industry Day", "sponsor_opportunity": "Industry track naming; showcase sponsor; Innovation Crawl"},
    {"day": "Day 3 — Wed Sept 16", "focus": "AI Innovation Day", "sponsor_opportunity": "AI Day presenting sponsor (HIGH VALUE) — ideal for tech companies"},
    {"day": "Day 4 — Thu Sept 17", "focus": "Investor Day", "sponsor_opportunity": "Investor Day presenting sponsor (HIGH VALUE) — ideal for finance/VC"},
    {"day": "Day 5 — Fri Sept 18", "focus": "Great MN Give Back Hack (w/ TCBAF)", "sponsor_opportunity": "Impact/CSR sponsor; hackathon naming rights"},
]

OUTPUT_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "research")

# ──────────────────────────────────────────────────────────────────────────────
# EXCEL GENERATION
# ──────────────────────────────────────────────────────────────────────────────

def hex_fill(hex_color):
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")

def header_font(bold=True, size=11, color=WHITE):
    return Font(bold=bold, size=size, color=color, name="Calibri")

def normal_font(bold=False, size=10, color=DARK_TEXT):
    return Font(bold=bold, size=size, color=color, name="Calibri")

def thin_border():
    thin = Side(border_style="thin", color="CCCCCC")
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def apply_header_row(ws, row, fill_hex, font_color=WHITE, height=20):
    ws.row_dimensions[row].height = height
    for cell in ws[row]:
        cell.fill = hex_fill(fill_hex)
        cell.font = Font(bold=True, size=10, color=font_color, name="Calibri")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border()

def auto_width(ws, min_w=10, max_w=50):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                v = str(cell.value) if cell.value else ""
                max_len = max(max_len, len(v))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_w, max(min_w, max_len + 2))

def build_excel(path):
    wb = openpyxl.Workbook()

    # ── Sheet 1: All Sponsors (Master) ───────────────────────────────────────
    ws1 = wb.active
    ws1.title = "All Sponsors — Master"
    ws1.sheet_view.showGridLines = False

    # Title banner
    ws1.merge_cells("A1:L1")
    title_cell = ws1["A1"]
    title_cell.value = "TWIN CITIES STARTUP WEEK — COMPLETE SPONSOR DATABASE"
    title_cell.fill = hex_fill(TCSW_TEAL)
    title_cell.font = Font(bold=True, size=14, color=WHITE, name="Calibri")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 28

    ws1.merge_cells("A2:L2")
    sub_cell = ws1["A2"]
    sub_cell.value = f"Synthesised from all research sources  •  All years 2014–2026  •  Generated {datetime.now().strftime('%B %d, %Y')}"
    sub_cell.fill = hex_fill(TCSW_GOLD)
    sub_cell.font = Font(bold=False, size=10, color=DARK_TEXT, name="Calibri")
    sub_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[2].height = 18

    headers = [
        "Year", "Organization", "Tier / Role", "Category", "Sector",
        "Key Contact", "Email", "Website", "Named Event", "Multi-Year?",
        "All Years Confirmed", "Notes / Activation Details"
    ]
    for col_idx, h in enumerate(headers, 1):
        c = ws1.cell(row=3, column=col_idx, value=h)
    apply_header_row(ws1, 3, TCSW_TEAL)

    # Sort by year desc, then org name
    sorted_sponsors = sorted(SPONSORS, key=lambda x: (-x["year"], x["organization"].lower()))

    alt_colors = [WHITE, TCSW_LIGHT]
    for row_idx, s in enumerate(sorted_sponsors, 4):
        row_data = [
            s["year"], s["organization"], s["tier_role"], s["category"],
            s["sector"], s["contact"], s["email"], s["website"],
            s["named_event"], "Yes" if s["multi_year"] else "No",
            s["years_confirmed"], s["notes"]
        ]
        color = alt_colors[row_idx % 2]
        for col_idx, val in enumerate(row_data, 1):
            c = ws1.cell(row=row_idx, column=col_idx, value=val)
            c.fill = hex_fill(color)
            c.font = normal_font()
            c.alignment = Alignment(vertical="top", wrap_text=(col_idx in [3, 5, 9, 12]))
            c.border = thin_border()
        ws1.row_dimensions[row_idx].height = 30

    auto_width(ws1)
    ws1.column_dimensions["B"].width = 32
    ws1.column_dimensions["C"].width = 30
    ws1.column_dimensions["I"].width = 35
    ws1.column_dimensions["L"].width = 48
    ws1.freeze_panes = "A4"

    # ── Sheet 2: 2026 Current Sponsors ───────────────────────────────────────
    ws2 = wb.create_sheet("2026 Sponsors")
    ws2.sheet_view.showGridLines = False

    ws2.merge_cells("A1:J1")
    c = ws2["A1"]
    c.value = "TCSW 2026 — CURRENT & CONFIRMED SPONSORS / PARTNERS"
    c.fill = hex_fill(TCSW_TEAL)
    c.font = Font(bold=True, size=13, color=WHITE, name="Calibri")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 26

    ws2.merge_cells("A2:J2")
    c2 = ws2["A2"]
    c2.value = "Event Dates: September 14–18, 2026  |  Contact: Angela.Eifert@BETA.MN  |  Subject: 'TCSW 2026 Sponsorship Prospectus Request'"
    c2.fill = hex_fill(TCSW_GOLD)
    c2.font = Font(size=9, color=DARK_TEXT, name="Calibri")
    c2.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[2].height = 16

    headers2 = ["Organization", "Tier / Role", "Category", "Sector", "Key Contact", "Email", "Website", "Named Event / Activation", "Multi-Year?", "Notes"]
    for ci, h in enumerate(headers2, 1):
        ws2.cell(row=3, column=ci, value=h)
    apply_header_row(ws2, 3, TCSW_TEAL)

    sponsors_2026 = [s for s in sorted_sponsors if s["year"] == 2026]
    for ri, s in enumerate(sponsors_2026, 4):
        row_data = [
            s["organization"], s["tier_role"], s["category"], s["sector"],
            s["contact"], s["email"], s["website"], s["named_event"],
            "Yes" if s["multi_year"] else "No", s["notes"]
        ]
        color = alt_colors[ri % 2]
        for ci, val in enumerate(row_data, 1):
            c = ws2.cell(row=ri, column=ci, value=val)
            c.fill = hex_fill(color)
            c.font = normal_font()
            c.alignment = Alignment(vertical="top", wrap_text=(ci in [2, 8, 10]))
            c.border = thin_border()
        ws2.row_dimensions[ri].height = 30

    auto_width(ws2)
    ws2.column_dimensions["A"].width = 35
    ws2.column_dimensions["B"].width = 32
    ws2.column_dimensions["H"].width = 40
    ws2.column_dimensions["J"].width = 45
    ws2.freeze_panes = "A4"

    # ── Sheet 3: By Year Summary ──────────────────────────────────────────────
    ws3 = wb.create_sheet("By Year Summary")
    ws3.sheet_view.showGridLines = False

    ws3.merge_cells("A1:F1")
    c = ws3["A1"]
    c.value = "TCSW SPONSORS — YEAR-BY-YEAR SUMMARY"
    c.fill = hex_fill(TCSW_TEAL)
    c.font = Font(bold=True, size=13, color=WHITE, name="Calibri")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws3.row_dimensions[1].height = 26

    from collections import defaultdict
    by_year = defaultdict(list)
    for s in SPONSORS:
        by_year[s["year"]].append(s)

    current_row = 3
    for year in sorted(by_year.keys(), reverse=True):
        year_sponsors = sorted(by_year[year], key=lambda x: x["organization"].lower())

        # Year header
        ws3.merge_cells(f"A{current_row}:F{current_row}")
        yc = ws3[f"A{current_row}"]
        yc.value = f"  {year}  —  {len(year_sponsors)} Sponsors / Partners"
        yc.fill = hex_fill(TCSW_GOLD)
        yc.font = Font(bold=True, size=11, color=DARK_TEXT, name="Calibri")
        yc.alignment = Alignment(vertical="center")
        ws3.row_dimensions[current_row].height = 22
        current_row += 1

        cols = ["Organization", "Tier / Role", "Category", "Key Contact", "Named Event", "Notes"]
        for ci, h in enumerate(cols, 1):
            c = ws3.cell(row=current_row, column=ci, value=h)
        apply_header_row(ws3, current_row, TCSW_MID, font_color=DARK_TEXT, height=16)
        current_row += 1

        for ri, s in enumerate(year_sponsors):
            row_data = [s["organization"], s["tier_role"], s["category"], s["contact"], s["named_event"], s["notes"]]
            color = alt_colors[ri % 2]
            for ci, val in enumerate(row_data, 1):
                c = ws3.cell(row=current_row, column=ci, value=val)
                c.fill = hex_fill(color)
                c.font = normal_font(size=9)
                c.alignment = Alignment(vertical="top", wrap_text=(ci in [2, 6]))
                c.border = thin_border()
            ws3.row_dimensions[current_row].height = 28
            current_row += 1

        current_row += 1

    auto_width(ws3)
    ws3.column_dimensions["A"].width = 35
    ws3.column_dimensions["B"].width = 30
    ws3.column_dimensions["F"].width = 45
    ws3.freeze_panes = "A3"

    # ── Sheet 4: 2026 Tiers & Event Structure ─────────────────────────────────
    ws4 = wb.create_sheet("2026 Tiers & Event")
    ws4.sheet_view.showGridLines = False

    ws4.merge_cells("A1:D1")
    c = ws4["A1"]
    c.value = "TCSW 2026 — SPONSORSHIP TIERS & EVENT STRUCTURE"
    c.fill = hex_fill(TCSW_TEAL)
    c.font = Font(bold=True, size=13, color=WHITE, name="Calibri")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws4.row_dimensions[1].height = 26

    ws4.merge_cells("A3:D3")
    ws4["A3"].value = "SPONSORSHIP TIERS"
    ws4["A3"].fill = hex_fill(TCSW_GOLD)
    ws4["A3"].font = Font(bold=True, size=11, color=DARK_TEXT, name="Calibri")
    ws4["A3"].alignment = Alignment(vertical="center")
    ws4.row_dimensions[3].height = 20

    tier_headers = ["Tier", "Minimum Investment", "Key Benefits"]
    for ci, h in enumerate(tier_headers, 1):
        ws4.cell(row=4, column=ci, value=h)
    apply_header_row(ws4, 4, TCSW_MID, font_color=DARK_TEXT)

    for ri, t in enumerate(TIERS_2026, 5):
        row_data = [t["tier"], t["min_investment"], t["benefits"]]
        color = alt_colors[ri % 2]
        for ci, val in enumerate(row_data, 1):
            c = ws4.cell(row=ri, column=ci, value=val)
            c.fill = hex_fill(color)
            c.font = normal_font()
            c.alignment = Alignment(vertical="top", wrap_text=(ci == 3))
            c.border = thin_border()
        ws4.row_dimensions[ri].height = 32

    current_row = len(TIERS_2026) + 7

    ws4.merge_cells(f"A{current_row}:D{current_row}")
    ws4[f"A{current_row}"].value = "2026 EVENT STRUCTURE & SPONSOR ACTIVATION OPPORTUNITIES"
    ws4[f"A{current_row}"].fill = hex_fill(TCSW_GOLD)
    ws4[f"A{current_row}"].font = Font(bold=True, size=11, color=DARK_TEXT, name="Calibri")
    ws4[f"A{current_row}"].alignment = Alignment(vertical="center")
    ws4.row_dimensions[current_row].height = 20
    current_row += 1

    day_headers = ["Day", "Theme / Focus", "Sponsor Activation Opportunity"]
    for ci, h in enumerate(day_headers, 1):
        ws4.cell(row=current_row, column=ci, value=h)
    apply_header_row(ws4, current_row, TCSW_MID, font_color=DARK_TEXT)
    current_row += 1

    for ri, d in enumerate(EVENT_DAYS):
        row_data = [d["day"], d["focus"], d["sponsor_opportunity"]]
        color = alt_colors[ri % 2]
        for ci, val in enumerate(row_data, 1):
            c = ws4.cell(row=current_row, column=ci, value=val)
            c.fill = hex_fill(color)
            c.font = normal_font()
            c.alignment = Alignment(vertical="top", wrap_text=(ci == 3))
            c.border = thin_border()
        ws4.row_dimensions[current_row].height = 32
        current_row += 1

    ws4.column_dimensions["A"].width = 28
    ws4.column_dimensions["B"].width = 30
    ws4.column_dimensions["C"].width = 60

    # ── Sheet 5: Key Contacts ─────────────────────────────────────────────────
    ws5 = wb.create_sheet("Key Contacts")
    ws5.sheet_view.showGridLines = False

    ws5.merge_cells("A1:D1")
    c = ws5["A1"]
    c.value = "TCSW — KEY CONTACTS & SOCIAL MEDIA"
    c.fill = hex_fill(TCSW_TEAL)
    c.font = Font(bold=True, size=13, color=WHITE, name="Calibri")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws5.row_dimensions[1].height = 26

    contact_headers = ["Name / Handle", "Title / Role", "Email", "Notes"]
    for ci, h in enumerate(contact_headers, 1):
        ws5.cell(row=3, column=ci, value=h)
    apply_header_row(ws5, 3, TCSW_TEAL)

    for ri, ct in enumerate(CONTACTS, 4):
        row_data = [ct["name"], ct["title"], ct["email"], ct["note"]]
        color = alt_colors[ri % 2]
        for ci, val in enumerate(row_data, 1):
            c = ws5.cell(row=ri, column=ci, value=val)
            c.fill = hex_fill(color)
            c.font = normal_font()
            c.alignment = Alignment(vertical="top")
            c.border = thin_border()

    current_row = len(CONTACTS) + 6
    ws5.merge_cells(f"A{current_row}:D{current_row}")
    ws5[f"A{current_row}"].value = "TCSW SOCIAL MEDIA"
    ws5[f"A{current_row}"].fill = hex_fill(TCSW_GOLD)
    ws5[f"A{current_row}"].font = Font(bold=True, size=11, color=DARK_TEXT, name="Calibri")
    ws5[f"A{current_row}"].alignment = Alignment(vertical="center")
    ws5.row_dimensions[current_row].height = 20
    current_row += 1

    social_headers = ["Platform", "Handle / URL", "", ""]
    for ci, h in enumerate(social_headers, 1):
        ws5.cell(row=current_row, column=ci, value=h)
    apply_header_row(ws5, current_row, TCSW_MID, font_color=DARK_TEXT)
    current_row += 1

    socials = [
        ("Twitter / X", "@tcstartupweek"),
        ("Instagram", "@tcstartupweek"),
        ("Facebook", "tcstartupweek"),
        ("BETA Twitter", "@betadotmn"),
        ("BETA Instagram", "@betadotmn"),
        ("LinkedIn (BETA)", "beta-mn"),
        ("Website", "https://tcstartupweek.com"),
        ("BETA Website", "https://www.beta.mn"),
    ]
    for ri, (platform, handle) in enumerate(socials):
        color = alt_colors[ri % 2]
        ws5.cell(row=current_row, column=1, value=platform).fill = hex_fill(color)
        ws5.cell(row=current_row, column=2, value=handle).fill = hex_fill(color)
        for ci in [1, 2]:
            ws5.cell(row=current_row, column=ci).font = normal_font()
            ws5.cell(row=current_row, column=ci).border = thin_border()
        current_row += 1

    auto_width(ws5)
    ws5.column_dimensions["A"].width = 30
    ws5.column_dimensions["B"].width = 40
    ws5.column_dimensions["D"].width = 45

    # ── Sheet 6: Multi-Year / Recurring Sponsors ─────────────────────────────
    ws6 = wb.create_sheet("Recurring Sponsors")
    ws6.sheet_view.showGridLines = False

    ws6.merge_cells("A1:F1")
    c = ws6["A1"]
    c.value = "TCSW — RECURRING / MULTI-YEAR SPONSORS"
    c.fill = hex_fill(TCSW_TEAL)
    c.font = Font(bold=True, size=13, color=WHITE, name="Calibri")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws6.row_dimensions[1].height = 26

    recurring = sorted([s for s in SPONSORS if s["multi_year"]], key=lambda x: x["organization"].lower())

    rec_headers = ["Organization", "Category", "Sector", "All Years", "Tier / Role", "Notes"]
    for ci, h in enumerate(rec_headers, 1):
        ws6.cell(row=3, column=ci, value=h)
    apply_header_row(ws6, 3, TCSW_TEAL)

    for ri, s in enumerate(recurring, 4):
        row_data = [s["organization"], s["category"], s["sector"], s["years_confirmed"], s["tier_role"], s["notes"]]
        color = alt_colors[ri % 2]
        for ci, val in enumerate(row_data, 1):
            c = ws6.cell(row=ri, column=ci, value=val)
            c.fill = hex_fill(color)
            c.font = normal_font()
            c.alignment = Alignment(vertical="top", wrap_text=(ci in [5, 6]))
            c.border = thin_border()
        ws6.row_dimensions[ri].height = 30

    auto_width(ws6)
    ws6.column_dimensions["A"].width = 35
    ws6.column_dimensions["E"].width = 32
    ws6.column_dimensions["F"].width = 48
    ws6.freeze_panes = "A4"

    wb.save(path)
    print(f"✅  Excel saved → {path}")


# ──────────────────────────────────────────────────────────────────────────────
# PDF GENERATION
# ──────────────────────────────────────────────────────────────────────────────

TCSW_TEAL_RGB   = colors.HexColor("#005F73")
TCSW_GOLD_RGB   = colors.HexColor("#EE9B00")
TCSW_LIGHT_RGB  = colors.HexColor("#E9F5F7")
TCSW_MID_RGB    = colors.HexColor("#94D2BD")
WHITE_RGB       = colors.white
DARK_RGB        = colors.HexColor("#001219")
GRAY_RGB        = colors.HexColor("#6C757D")

def get_styles():
    styles = getSampleStyleSheet()
    custom = {}

    custom["cover_title"] = ParagraphStyle(
        "cover_title", parent=styles["Title"],
        fontSize=28, textColor=WHITE_RGB, alignment=TA_CENTER,
        fontName="Helvetica-Bold", spaceAfter=6
    )
    custom["cover_sub"] = ParagraphStyle(
        "cover_sub", parent=styles["Normal"],
        fontSize=13, textColor=TCSW_GOLD_RGB, alignment=TA_CENTER,
        fontName="Helvetica", spaceAfter=4
    )
    custom["cover_meta"] = ParagraphStyle(
        "cover_meta", parent=styles["Normal"],
        fontSize=10, textColor=WHITE_RGB, alignment=TA_CENTER,
        fontName="Helvetica"
    )
    custom["section_title"] = ParagraphStyle(
        "section_title", parent=styles["Heading1"],
        fontSize=14, textColor=WHITE_RGB, alignment=TA_LEFT,
        fontName="Helvetica-Bold", spaceAfter=6, spaceBefore=16,
        backColor=TCSW_TEAL_RGB, leftIndent=-6, borderPad=4
    )
    custom["year_title"] = ParagraphStyle(
        "year_title", parent=styles["Heading2"],
        fontSize=12, textColor=DARK_RGB, alignment=TA_LEFT,
        fontName="Helvetica-Bold", spaceAfter=4, spaceBefore=12,
        backColor=TCSW_GOLD_RGB, leftIndent=-4, borderPad=3
    )
    custom["body"] = ParagraphStyle(
        "body", parent=styles["Normal"],
        fontSize=9, textColor=DARK_RGB, fontName="Helvetica",
        spaceAfter=2, leading=13
    )
    custom["body_bold"] = ParagraphStyle(
        "body_bold", parent=styles["Normal"],
        fontSize=9, textColor=DARK_RGB, fontName="Helvetica-Bold",
        spaceAfter=2
    )
    custom["small"] = ParagraphStyle(
        "small", parent=styles["Normal"],
        fontSize=8, textColor=GRAY_RGB, fontName="Helvetica",
        spaceAfter=1
    )
    custom["table_header"] = ParagraphStyle(
        "table_header", parent=styles["Normal"],
        fontSize=8, textColor=WHITE_RGB, fontName="Helvetica-Bold",
        alignment=TA_CENTER
    )
    custom["table_cell"] = ParagraphStyle(
        "table_cell", parent=styles["Normal"],
        fontSize=8, textColor=DARK_RGB, fontName="Helvetica",
        leading=11
    )
    return custom

def make_table_style(header_color=TCSW_TEAL_RGB, alt_color=TCSW_LIGHT_RGB, n_rows=1):
    style = [
        ("BACKGROUND", (0, 0), (-1, 0), header_color),
        ("TEXTCOLOR", (0, 0), (-1, 0), WHITE_RGB),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 8),
        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 1), (-1, -1), 8),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [WHITE_RGB, alt_color]),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#CCCCCC")),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ]
    return TableStyle(style)

def build_pdf(path):
    doc = SimpleDocTemplate(
        path,
        pagesize=letter,
        leftMargin=0.6*inch, rightMargin=0.6*inch,
        topMargin=0.6*inch, bottomMargin=0.7*inch,
        title="TCSW Sponsors Master Report",
        author="TCSW Research",
        subject="Twin Cities Startup Week — Complete Sponsor Database"
    )

    styles = get_styles()
    story = []
    W = letter[0] - 1.2*inch  # usable width

    def section_hdr(text):
        return [
            Spacer(1, 8),
            Table([[Paragraph(text, styles["section_title"])]],
                  colWidths=[W],
                  style=TableStyle([
                      ("BACKGROUND", (0, 0), (-1, -1), TCSW_TEAL_RGB),
                      ("LEFTPADDING", (0, 0), (-1, -1), 8),
                      ("TOPPADDING", (0, 0), (-1, -1), 6),
                      ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                  ])),
            Spacer(1, 6),
        ]

    def year_hdr(text):
        return [
            Spacer(1, 6),
            Table([[Paragraph(text, styles["year_title"])]],
                  colWidths=[W],
                  style=TableStyle([
                      ("BACKGROUND", (0, 0), (-1, -1), TCSW_GOLD_RGB),
                      ("LEFTPADDING", (0, 0), (-1, -1), 8),
                      ("TOPPADDING", (0, 0), (-1, -1), 4),
                      ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                  ])),
            Spacer(1, 4),
        ]

    # ── Cover ─────────────────────────────────────────────────────────────────
    cover_table = Table(
        [[Paragraph("TWIN CITIES STARTUP WEEK", styles["cover_title"])],
         [Paragraph("Complete Sponsor Database", styles["cover_sub"])],
         [Spacer(1, 10)],
         [Paragraph("All Years 2014 – 2026  •  All Tiers & Categories", styles["cover_meta"])],
         [Paragraph(f"Synthesised from all research sources  •  Generated {datetime.now().strftime('%B %d, %Y')}", styles["cover_meta"])],
         [Spacer(1, 6)],
         [Paragraph("Organizer: BETA.MN (The Beta Group)  |  Contact: Angela.Eifert@BETA.MN", styles["cover_meta"])],
        ],
        colWidths=[W],
        style=TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), TCSW_TEAL_RGB),
            ("TOPPADDING", (0, 0), (-1, -1), 10),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
            ("LEFTPADDING", (0, 0), (-1, -1), 20),
            ("RIGHTPADDING", (0, 0), (-1, -1), 20),
            ("ROUNDEDCORNERS", [6]),
        ])
    )
    story.append(cover_table)
    story.append(Spacer(1, 14))

    # Stats summary
    total_orgs = len(set(s["organization"] for s in SPONSORS))
    years_covered = sorted(set(s["year"] for s in SPONSORS))
    multi_year_count = len([s for s in SPONSORS if s["multi_year"]])
    sponsors_2026_count = len([s for s in SPONSORS if s["year"] == 2026])

    stat_data = [
        ["Total Organizations", "Years of Data", "Multi-Year Sponsors", "2026 Confirmed"],
        [str(total_orgs), f"2014–2026 ({len(years_covered)} years)", str(multi_year_count), str(sponsors_2026_count)],
    ]
    stat_table = Table(stat_data, colWidths=[W/4]*4,
                       style=TableStyle([
                           ("BACKGROUND", (0, 0), (-1, 0), TCSW_TEAL_RGB),
                           ("BACKGROUND", (0, 1), (-1, 1), TCSW_LIGHT_RGB),
                           ("TEXTCOLOR", (0, 0), (-1, 0), WHITE_RGB),
                           ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                           ("FONTNAME", (0, 1), (-1, 1), "Helvetica-Bold"),
                           ("FONTSIZE", (0, 0), (-1, -1), 10),
                           ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                           ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                           ("TOPPADDING", (0, 0), (-1, -1), 8),
                           ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
                           ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#AAAAAA")),
                       ]))
    story.append(stat_table)
    story.append(Spacer(1, 12))

    # ── Section 1: 2026 Sponsors ─────────────────────────────────────────────
    story.extend(section_hdr("SECTION 1 — 2026 CURRENT & CONFIRMED SPONSORS"))

    story.append(Paragraph(
        "The following organizations are confirmed sponsors, partners, or co-hosts for TCSW 2026 "
        "(September 14–18, 2026). The event spans five days across six campuses. "
        "Sponsorship tiers range from Builder ($5,000+) through Presenter, Host, and Premier. "
        "Contact: <b>Angela.Eifert@BETA.MN</b>  |  Subject: 'TCSW 2026 Sponsorship Prospectus Request'",
        styles["body"]
    ))
    story.append(Spacer(1, 6))

    sponsors_2026 = [s for s in sorted(SPONSORS, key=lambda x: x["organization"].lower()) if s["year"] == 2026]

    tbl_data_2026 = [["Organization", "Tier / Role", "Sector", "Key Contact", "Named Event / Notes"]]
    for s in sponsors_2026:
        contact_str = s["contact"] if s["contact"] else "—"
        event_str = s["named_event"] if s["named_event"] else s["notes"][:60] if s["notes"] else "—"
        tbl_data_2026.append([
            Paragraph(s["organization"], styles["table_cell"]),
            Paragraph(s["tier_role"], styles["table_cell"]),
            Paragraph(s["sector"], styles["table_cell"]),
            Paragraph(contact_str, styles["table_cell"]),
            Paragraph(event_str, styles["table_cell"]),
        ])

    col_w = [W * r for r in [0.23, 0.22, 0.14, 0.19, 0.22]]
    t = Table(tbl_data_2026, colWidths=col_w, repeatRows=1)
    t.setStyle(make_table_style())
    story.append(t)
    story.append(Spacer(1, 8))

    # 2026 Event days
    story.extend(year_hdr("2026 Event Schedule & Activation Opportunities"))
    day_data = [["Day", "Theme / Focus", "Sponsor Opportunity"]]
    for d in EVENT_DAYS:
        day_data.append([
            Paragraph(d["day"], styles["table_cell"]),
            Paragraph(d["focus"], styles["table_cell"]),
            Paragraph(d["sponsor_opportunity"], styles["table_cell"]),
        ])
    t2 = Table(day_data, colWidths=[W*0.22, W*0.25, W*0.53], repeatRows=1)
    t2.setStyle(make_table_style(header_color=TCSW_TEAL_RGB))
    story.append(t2)
    story.append(Spacer(1, 8))

    # 2026 Tiers
    story.extend(year_hdr("2026 Sponsorship Tiers"))
    tier_data = [["Tier", "Min. Investment", "Key Benefits"]]
    for t_item in TIERS_2026:
        tier_data.append([
            Paragraph(f"<b>{t_item['tier']}</b>", styles["table_cell"]),
            Paragraph(t_item["min_investment"], styles["table_cell"]),
            Paragraph(t_item["benefits"], styles["table_cell"]),
        ])
    tt = Table(tier_data, colWidths=[W*0.14, W*0.16, W*0.70], repeatRows=1)
    tt.setStyle(make_table_style(header_color=TCSW_TEAL_RGB))
    story.append(tt)

    story.append(PageBreak())

    # ── Section 2: By Year ────────────────────────────────────────────────────
    story.extend(section_hdr("SECTION 2 — SPONSORS BY YEAR (2025 → 2014)"))

    from collections import defaultdict
    by_year = defaultdict(list)
    for s in SPONSORS:
        if s["year"] < 2026:
            by_year[s["year"]].append(s)

    for year in sorted(by_year.keys(), reverse=True):
        year_sponsors = sorted(by_year[year], key=lambda x: x["organization"].lower())
        story.extend(year_hdr(f"{year}  —  {len(year_sponsors)} Sponsors / Partners Confirmed"))

        y_data = [["Organization", "Tier / Role", "Sector / Category", "Named Event / Activation", "Notes"]]
        for s in year_sponsors:
            event_str = s["named_event"] if s["named_event"] else "—"
            y_data.append([
                Paragraph(s["organization"], styles["table_cell"]),
                Paragraph(s["tier_role"], styles["table_cell"]),
                Paragraph(s["sector"], styles["table_cell"]),
                Paragraph(event_str, styles["table_cell"]),
                Paragraph(s["notes"][:70] if s["notes"] else "", styles["table_cell"]),
            ])

        col_widths = [W * r for r in [0.22, 0.20, 0.15, 0.22, 0.21]]
        yt = Table(y_data, colWidths=col_widths, repeatRows=1)
        yt.setStyle(make_table_style(header_color=TCSW_MID_RGB, alt_color=TCSW_LIGHT_RGB))
        story.append(yt)
        story.append(Spacer(1, 4))

    story.append(PageBreak())

    # ── Section 3: Recurring / Multi-Year ────────────────────────────────────
    story.extend(section_hdr("SECTION 3 — RECURRING & MULTI-YEAR SPONSORS"))
    story.append(Paragraph(
        "These organizations have supported Twin Cities Startup Week across multiple years, "
        "demonstrating sustained commitment to the MN startup ecosystem.",
        styles["body"]
    ))
    story.append(Spacer(1, 6))

    recurring = sorted([s for s in SPONSORS if s["multi_year"]], key=lambda x: x["organization"].lower())
    rec_data = [["Organization", "Sector", "Years Active", "Highest Tier / Role", "Notes"]]
    for s in recurring:
        rec_data.append([
            Paragraph(f"<b>{s['organization']}</b>", styles["table_cell"]),
            Paragraph(s["sector"], styles["table_cell"]),
            Paragraph(s["years_confirmed"], styles["table_cell"]),
            Paragraph(s["tier_role"], styles["table_cell"]),
            Paragraph(s["notes"][:65] if s["notes"] else "", styles["table_cell"]),
        ])

    rt = Table(rec_data, colWidths=[W * r for r in [0.24, 0.16, 0.14, 0.22, 0.24]], repeatRows=1)
    rt.setStyle(make_table_style())
    story.append(rt)
    story.append(PageBreak())

    # ── Section 4: Fortune 500 Spotlight ─────────────────────────────────────
    story.extend(section_hdr("SECTION 4 — FORTUNE 500 SPONSOR SPOTLIGHT (2019)"))
    story.append(Paragraph(
        "2019 marked TCSW's first year with Fortune 500 title sponsors. Three major Minnesota-based "
        "corporations served as title sponsors alongside Microsoft and AWS, representing the event's "
        "peak corporate sponsorship engagement to date.",
        styles["body"]
    ))
    story.append(Spacer(1, 8))

    fortune500 = [
        {
            "company": "Target Corporation",
            "role": "Title Sponsor 2019",
            "contact": "Gene Han, VP New Ventures & Accelerators",
            "activation": "Hosted Kickoff Event + Retail Track at Target Plaza Commons (Minneapolis HQ)",
            "quote": '"Startups and corporations can work together in a way that is mutually beneficial. We embrace working with startups through our portfolio of accelerator programs."',
            "why": "Community investment, corporate innovation pipeline, talent",
        },
        {
            "company": "Cargill",
            "role": "Title Sponsor 2019 (also co-sponsor 2018)",
            "contact": "Lawrence Wang, Digitalization & Analytics Strategy Lead",
            "activation": "Co-sponsored Techstars Farm to Fork; Demo Day October 15",
            "quote": '"The pace of innovation is so fast that close-knit relationships with early stage founders is required to remain ahead of the competition."',
            "why": "AgTech + Food Technology track; startup partnership pipeline; Wayzata, MN HQ",
        },
        {
            "company": "3M",
            "role": "Title Sponsor 2019 (also sponsor 2018)",
            "contact": "Ben Wright, Director of 3M Ventures",
            "activation": "BETA Backers Program (investors meet early-stage MN startups)",
            "quote": '"Twin Cities Startup Week is a great opportunity to collaborate with the local entrepreneurial community. Innovation is at the heart of 3M\'s culture."',
            "why": "3M Ventures deal flow access; MN Fortune 500; advanced manufacturing / AI Day fit",
        },
    ]

    for company in fortune500:
        c_data = [
            [Paragraph(f"<b>{company['company']}</b>  —  {company['role']}", styles["body_bold"])],
            [Paragraph(f"<b>Contact (2019):</b> {company['contact']}", styles["body"])],
            [Paragraph(f"<b>Activation:</b> {company['activation']}", styles["body"])],
            [Paragraph(f"<b>Why they sponsor:</b> {company['why']}", styles["body"])],
            [Paragraph(f"<i>{company['quote']}</i>", styles["small"])],
        ]
        ct = Table(c_data, colWidths=[W],
                   style=TableStyle([
                       ("BACKGROUND", (0, 0), (-1, -1), TCSW_LIGHT_RGB),
                       ("BACKGROUND", (0, 0), (-1, 0), TCSW_MID_RGB),
                       ("LEFTPADDING", (0, 0), (-1, -1), 10),
                       ("RIGHTPADDING", (0, 0), (-1, -1), 10),
                       ("TOPPADDING", (0, 0), (-1, -1), 4),
                       ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                       ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#CCCCCC")),
                   ]))
        story.append(ct)
        story.append(Spacer(1, 8))

    story.append(PageBreak())

    # ── Section 5: Full Master Table ─────────────────────────────────────────
    story.extend(section_hdr("SECTION 5 — COMPLETE MASTER SPONSOR TABLE (ALL YEARS)"))
    story.append(Paragraph(
        "Every confirmed sponsor, partner, venue, and co-host from all available research, "
        "sorted by year (most recent first).",
        styles["body"]
    ))
    story.append(Spacer(1, 6))

    all_sorted = sorted(SPONSORS, key=lambda x: (-x["year"], x["organization"].lower()))
    master_data = [["Yr", "Organization", "Tier / Role", "Sector", "Key Contact", "Multi-Yr?"]]
    for s in all_sorted:
        master_data.append([
            str(s["year"]),
            Paragraph(s["organization"], styles["table_cell"]),
            Paragraph(s["tier_role"], styles["table_cell"]),
            Paragraph(s["sector"], styles["table_cell"]),
            Paragraph(s["contact"][:35] if s["contact"] else "—", styles["table_cell"]),
            "Yes" if s["multi_year"] else "No",
        ])

    mt = Table(master_data, colWidths=[W * r for r in [0.05, 0.26, 0.22, 0.15, 0.22, 0.10]], repeatRows=1)
    mt.setStyle(make_table_style())
    story.append(mt)
    story.append(Spacer(1, 12))

    # Footer note
    story.append(HRFlowable(width=W, thickness=1, color=TCSW_TEAL_RGB))
    story.append(Spacer(1, 4))
    story.append(Paragraph(
        f"Twin Cities Startup Week — Sponsor Research Report  •  "
        f"Generated {datetime.now().strftime('%B %d, %Y')}  •  "
        "Sources: tcstartupweek.com, beta.mn, start-midwest.com, tcsw2023.sched.com, tcbmag.com, medium.com/twin-cities-startup-week",
        styles["small"]
    ))

    doc.build(story)
    print(f"✅  PDF saved  → {path}")


# ──────────────────────────────────────────────────────────────────────────────
# MAIN
# ──────────────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    excel_path = os.path.join(OUTPUT_DIR, "TCSW_Sponsors_Master.xlsx")
    pdf_path   = os.path.join(OUTPUT_DIR, "TCSW_Sponsors_Master.pdf")

    print("Building Excel workbook…")
    build_excel(excel_path)

    print("Building PDF report…")
    build_pdf(pdf_path)

    print("\nDone! Files saved to:")
    print(f"  📊  {excel_path}")
    print(f"  📄  {pdf_path}")
