"""
build_master_docs.py
====================
Generates the definitive TCSW_Sponsors_Master.xlsx and TCSW_Sponsors_Master.pdf
with all sponsors, partners, venues — including 21 previously missing entries.
"""

import csv
from pathlib import Path
from datetime import date

# ── Paths ──────────────────────────────────────────────────────────────────────
REPO = Path(__file__).resolve().parent.parent
OUT  = REPO / "research"

TODAY = date.today().strftime("%B %d, %Y")

# ══════════════════════════════════════════════════════════════════════════════
# MASTER DATA — every confirmed sponsor, partner, venue, co-host
# Fields: year, organization, tier, category, sector, key_contact, email,
#         website, named_event, multi_year, ad_placement, notes
# ══════════════════════════════════════════════════════════════════════════════

MASTER = [
    # ── 2026 ─────────────────────────────────────────────────────────────────
    {"year": 2026, "organization": "BETA.MN (The Beta Group)", "tier": "Organizer / Operator", "category": "Nonprofit / Community", "sector": "Nonprofit", "key_contact": "Angela Eifert — Executive Director", "email": "Angela.Eifert@BETA.MN", "website": "https://www.beta.mn", "named_event": "", "multi_year": "Yes", "ad_placement": "All materials", "notes": "501(c)(3) nonprofit; primary operator of TCSW every year"},
    {"year": 2026, "organization": "Communiful", "tier": "Co-Host / Co-Organizer", "category": "Coworking / Creative Hub", "sector": "Tech / Community", "key_contact": "Hayley Brooks — Founder", "email": "", "website": "https://communiful.com", "named_event": "", "multi_year": "Yes", "ad_placement": "Website + signage", "notes": "Creative hub & coworking; co-hosting 2025 & 2026"},
    {"year": 2026, "organization": "JPMorgan Chase", "tier": "Premier Sponsor", "category": "Financial Services", "sector": "Finance / Banking", "key_contact": "", "email": "", "website": "https://jpmorganchase.com", "named_event": "", "multi_year": "Yes", "ad_placement": "Website + stage", "notes": "BETA Showcase sponsor 2023; listed 2026 sponsor"},
    {"year": 2026, "organization": "Augeo", "tier": "Sponsor", "category": "Technology / Loyalty", "sector": "Technology", "key_contact": "", "email": "", "website": "https://augeo.com", "named_event": "", "multi_year": "No", "ad_placement": "Website", "notes": "Listed 2026 sponsor"},
    {"year": 2026, "organization": "Biometrica Health", "tier": "Sponsor", "category": "Health Tech", "sector": "Healthcare Technology", "key_contact": "", "email": "", "website": "", "named_event": "", "multi_year": "No", "ad_placement": "Website", "notes": "Listed 2026 sponsor"},
    {"year": 2026, "organization": "Bread & Butter Ventures", "tier": "Ecosystem Partner / VC Panel", "category": "Venture Capital", "sector": "VC / Investor", "key_contact": "", "email": "", "website": "https://breadandbutter.vc", "named_event": "VC Panel", "multi_year": "Yes", "ad_placement": "Website + panel", "notes": "MN VC; listed 2026 sponsor"},
    {"year": 2026, "organization": "Brooksource", "tier": "Sponsor", "category": "Staffing / Technology", "sector": "Professional Services", "key_contact": "", "email": "", "website": "https://brooksource.com", "named_event": "", "multi_year": "No", "ad_placement": "Website", "notes": "Listed 2026 sponsor"},
    {"year": 2026, "organization": "ezer21", "tier": "Sponsor", "category": "Technology / Startup", "sector": "Technology", "key_contact": "", "email": "", "website": "", "named_event": "", "multi_year": "No", "ad_placement": "Website", "notes": "Listed 2026 sponsor"},
    {"year": 2026, "organization": "Full Stack Saint Paul Scale Up", "tier": "Sponsor / Program", "category": "Ecosystem / Program", "sector": "Startup Support", "key_contact": "", "email": "", "website": "", "named_event": "", "multi_year": "No", "ad_placement": "Website", "notes": "Scale-up program; listed 2026 sponsor"},
    {"year": 2026, "organization": "Groove Capital", "tier": "Event Partner — Pitch Night", "category": "Venture Capital", "sector": "VC / Investor", "key_contact": "", "email": "", "website": "https://groovecapital.com", "named_event": "Pitch Night (portfolio showcase)", "multi_year": "Yes", "ad_placement": "Website + event", "notes": "MN-based VC; hosts portfolio showcase pitch night"},
    {"year": 2026, "organization": "Idea Fund", "tier": "Sponsor / Happy Hour", "category": "Venture Capital", "sector": "Regional VC (La Crosse, WI)", "key_contact": "", "email": "", "website": "", "named_event": "Monday Happy Hour", "multi_year": "Yes", "ad_placement": "Event", "notes": "Regional VC; hosted Monday happy hour 2025"},
    {"year": 2026, "organization": "Launch Minnesota", "tier": "Government Partner", "category": "Government", "sector": "Government / Economic Dev", "key_contact": "", "email": "", "website": "https://launchmn.com", "named_event": "Tuesday Happy Hour", "multi_year": "Yes", "ad_placement": "Website + event", "notes": "State of MN startup initiative; hosts happy hour annually"},
    {"year": 2026, "organization": "Mairs & Power", "tier": "Sponsor / Panel", "category": "Financial Services", "sector": "Investment Management", "key_contact": "Scott Burns — Managing Director (Venture Capital)", "email": "", "website": "https://mairsandpower.com", "named_event": "B2B Scaling Panel (Thursday)", "multi_year": "Yes", "ad_placement": "Website + panel", "notes": "Scott Burns spoke on Thursday B2B scaling panel in 2025"},
    {"year": 2026, "organization": "MN Cup / University of Minnesota", "tier": "Competition Partner", "category": "Education / Competition", "sector": "Competition / University", "key_contact": "", "email": "hellomncup@umn.edu", "website": "https://mncup.org", "named_event": "Grand Finale ($400K+ prizes)", "multi_year": "Yes", "ad_placement": "All materials", "notes": "Annual pitch competition Grand Finale held during TCSW every year"},
    {"year": 2026, "organization": "Northeast Bank", "tier": "Sponsor", "category": "Financial Services", "sector": "Banking / Finance", "key_contact": "", "email": "", "website": "https://northeastbank.com", "named_event": "", "multi_year": "No", "ad_placement": "Website", "notes": "Listed 2026 sponsor"},
    {"year": 2026, "organization": "Shakopee Mdewakanton Sioux Community", "tier": "Sponsor", "category": "Government / Tribal", "sector": "Tribal Nation / Community", "key_contact": "", "email": "", "website": "https://shakopeedakota.org", "named_event": "", "multi_year": "No", "ad_placement": "Website", "notes": "Tribal nation sponsor — MN community partner"},
    {"year": 2026, "organization": "Taft Law", "tier": "Sponsor — Legal", "category": "Professional Services", "sector": "Legal / Law Firm", "key_contact": "", "email": "", "website": "https://taftlaw.com", "named_event": "Exit Planning Sessions", "multi_year": "No", "ad_placement": "Website + session", "notes": "Annual exit planning legal sessions"},
    # ── 2025 ─────────────────────────────────────────────────────────────────
    {"year": 2025, "organization": "BETA.MN (The Beta Group)", "tier": "Organizer / Operator", "category": "Nonprofit / Community", "sector": "Nonprofit", "key_contact": "Angela Eifert — Executive Director", "email": "Angela.Eifert@BETA.MN", "website": "https://www.beta.mn", "named_event": "", "multi_year": "Yes", "ad_placement": "All materials", "notes": "Primary organizer"},
    {"year": 2025, "organization": "Communiful", "tier": "Co-Host / Co-Organizer", "category": "Coworking / Creative Hub", "sector": "Tech / Community", "key_contact": "Hayley Brooks — Founder", "email": "", "website": "https://communiful.com", "named_event": "", "multi_year": "Yes", "ad_placement": "Website + signage", "notes": "Co-hosting with BETA.MN"},
    {"year": 2025, "organization": "DEED (MN Dept of Employment & Economic Dev)", "tier": "Government Partner — Resource Fair", "category": "Government", "sector": "Government", "key_contact": "", "email": "", "website": "https://mn.gov/deed", "named_event": "Thursday St. Paul Resource Fair", "multi_year": "Yes", "ad_placement": "Event", "notes": "State agency providing startup resources"},
    {"year": 2025, "organization": "EmpowerHER", "tier": "Community Partner — Investor Networking", "category": "Community / DEI", "sector": "Organization", "key_contact": "", "email": "", "website": "", "named_event": "Investor-Founder Networking Event (Tuesday)", "multi_year": "No", "ad_placement": "Event", "notes": "Women founders program; co-hosted Tuesday networking"},
    {"year": 2025, "organization": "gener8tor", "tier": "Event Partner — Showcase", "category": "Startup Accelerator", "sector": "Accelerator", "key_contact": "", "email": "", "website": "https://gener8tor.com", "named_event": "BETA × gener8tor Showcase (200+ attendees)", "multi_year": "No", "ad_placement": "Event + website", "notes": "Startup accelerator; co-hosted showcase with 200+ attendees"},
    {"year": 2025, "organization": "Great North Ventures", "tier": "Event Partner — Investor Networking", "category": "Venture Capital", "sector": "VC / Investor", "key_contact": "", "email": "", "website": "https://greatnorthventures.com", "named_event": "Investor-Founder Networking Event (Tuesday)", "multi_year": "No", "ad_placement": "Event", "notes": "MN VC; co-hosted Tuesday investor-founder networking"},
    {"year": 2025, "organization": "Hyde Park Venture Partners", "tier": "VC Panel Participant", "category": "Venture Capital", "sector": "VC / Investor", "key_contact": "", "email": "", "website": "https://hydeparkvp.com", "named_event": "Monday VC Panel", "multi_year": "No", "ad_placement": "Panel", "notes": "Participated in Monday VC panel"},
    {"year": 2025, "organization": "M25", "tier": "VC Panel Participant", "category": "Venture Capital", "sector": "VC / Investor (Chicago)", "key_contact": "", "email": "", "website": "https://m25.vc", "named_event": "Monday VC Panel", "multi_year": "No", "ad_placement": "Panel", "notes": "Chicago-based VC; participated in Monday VC panel"},
    {"year": 2025, "organization": "Matchstick Ventures", "tier": "VC Panel Participant", "category": "Venture Capital", "sector": "VC / Investor", "key_contact": "", "email": "", "website": "https://matchstickventures.com", "named_event": "Monday VC Panel", "multi_year": "No", "ad_placement": "Panel", "notes": "MN VC; participated in Monday VC panel"},
    {"year": 2025, "organization": "MN:EVC", "tier": "Ecosystem Partner — Networking", "category": "Venture Capital", "sector": "Investor Network", "key_contact": "", "email": "", "website": "", "named_event": "Networking Event", "multi_year": "No", "ad_placement": "Event", "notes": "Minnesota Early-stage Venture Capital network"},
    {"year": 2025, "organization": "Naturally MN", "tier": "Event Partner — Food & Bev Showcase", "category": "Food & Beverage", "sector": "Food & Bev", "key_contact": "", "email": "", "website": "", "named_event": "Food & Bev Showcase", "multi_year": "No", "ad_placement": "Event", "notes": "MN food & beverage startup showcase"},
    {"year": 2025, "organization": "Bread & Butter Ventures", "tier": "Ecosystem Partner", "category": "Venture Capital", "sector": "VC / Investor", "key_contact": "", "email": "", "website": "https://breadandbutter.vc", "named_event": "", "multi_year": "Yes", "ad_placement": "Website", "notes": "MN VC ecosystem partner"},
    {"year": 2025, "organization": "Groove Capital", "tier": "Event Partner", "category": "Venture Capital", "sector": "VC / Investor", "key_contact": "", "email": "", "website": "https://groovecapital.com", "named_event": "", "multi_year": "Yes", "ad_placement": "Website", "notes": "MN VC ecosystem partner"},
    # ── 2024 ─────────────────────────────────────────────────────────────────
    {"year": 2024, "organization": "BETA.MN (The Beta Group)", "tier": "Organizer / Operator", "category": "Nonprofit / Community", "sector": "Nonprofit", "key_contact": "Angela Eifert — Executive Director", "email": "Angela.Eifert@BETA.MN", "website": "https://www.beta.mn", "named_event": "", "multi_year": "Yes", "ad_placement": "All materials", "notes": "Primary organizer"},
    # ── 2023 ─────────────────────────────────────────────────────────────────
    {"year": 2023, "organization": "BETA.MN (The Beta Group)", "tier": "Organizer / Operator", "category": "Nonprofit / Community", "sector": "Nonprofit", "key_contact": "Angela Eifert — Executive Director", "email": "Angela.Eifert@BETA.MN", "website": "https://www.beta.mn", "named_event": "", "multi_year": "Yes", "ad_placement": "All materials", "notes": "Primary organizer"},
    {"year": 2023, "organization": "Optum (UnitedHealth Group)", "tier": "Title Sponsor — Cup Naming Rights", "category": "Health Tech", "sector": "Healthcare / Insurance", "key_contact": "", "email": "", "website": "https://optum.com", "named_event": "Optum Cup (pitch competition naming rights)", "multi_year": "No", "ad_placement": "All materials + named event", "notes": "Named the MN Cup pitch competition 'Optum Cup'"},
    {"year": 2023, "organization": "TriNet", "tier": "Presenting Sponsor — VC Fireside Chat", "category": "HR / Benefits", "sector": "HR / Benefits", "key_contact": "", "email": "", "website": "https://trinet.com", "named_event": "VC Fireside Chat + WeWork Mixer", "multi_year": "No", "ad_placement": "Stage + branding", "notes": "Presenting sponsor; branding on stage; WeWork mixer co-host"},
    {"year": 2023, "organization": "Twin Ignition", "tier": "Presenting Sponsor — Opening Party", "category": "Ecosystem / Community", "sector": "Ecosystem / Community", "key_contact": "", "email": "", "website": "https://twinignition.com", "named_event": "Opening Party Presented by Twin Ignition", "multi_year": "Yes", "ad_placement": "Named event + website", "notes": "Co-hosted opening party; provided Startup Garage + Keg House venue"},
    {"year": 2023, "organization": "Workbox", "tier": "Venue Sponsor", "category": "Coworking / Real Estate", "sector": "Coworking / Real Estate", "key_contact": "", "email": "", "website": "https://workbox.com", "named_event": "", "multi_year": "Yes", "ad_placement": "Website + signage", "notes": "Premium coworking space; hosted multiple TCSW sessions"},
    {"year": 2023, "organization": "FINNOVATION Lab", "tier": "Venue Sponsor", "category": "Fintech", "sector": "Fintech", "key_contact": "", "email": "", "website": "https://finnovationlab.com", "named_event": "", "multi_year": "Yes", "ad_placement": "Website + signage", "notes": "Training Lab venue"},
    {"year": 2023, "organization": "Luminary Arts Center", "tier": "Venue Partner — Main Stage", "category": "Arts / Event Venue", "sector": "Venue", "key_contact": "", "email": "", "website": "https://luminaryartscenter.com", "named_event": "", "multi_year": "No", "ad_placement": "Signage", "notes": "Main stage + breakout rooms 2023"},
    {"year": 2023, "organization": "Augurian & Cloudburst SBC", "tier": "Venue Sponsor", "category": "Marketing / Agency", "sector": "Marketing", "key_contact": "", "email": "", "website": "https://augurian.com", "named_event": "Cloudburst SBC venue", "multi_year": "No", "ad_placement": "Venue branding", "notes": "Co-branded venue for Cloudburst SBC sessions"},
    {"year": 2023, "organization": "Lab651 HQ", "tier": "Venue Sponsor", "category": "Technology / Software", "sector": "Technology", "key_contact": "", "email": "", "website": "https://lab651.com", "named_event": "", "multi_year": "No", "ad_placement": "Venue branding", "notes": "AI + software session host"},
    {"year": 2023, "organization": "Stratasys", "tier": "Venue / Session Sponsor", "category": "Manufacturing / 3D Printing", "sector": "Manufacturing / Tech", "key_contact": "", "email": "", "website": "https://stratasys.com", "named_event": "", "multi_year": "No", "ad_placement": "Venue branding", "notes": "3D printing session host"},
    {"year": 2023, "organization": "West Monroe", "tier": "Venue / Session Sponsor", "category": "Consulting / Technology", "sector": "Consulting", "key_contact": "", "email": "", "website": "https://westmonroe.com", "named_event": "", "multi_year": "No", "ad_placement": "Venue branding", "notes": "AI Day session host"},
    {"year": 2023, "organization": "JPMorgan Chase", "tier": "BETA Showcase Sponsor", "category": "Financial Services", "sector": "Finance / Banking", "key_contact": "", "email": "", "website": "https://jpmorganchase.com", "named_event": "BETA Fall Showcase", "multi_year": "Yes", "ad_placement": "Event + website", "notes": "Sponsored BETA Showcase 2023"},
    {"year": 2023, "organization": "WeWork (NorthLoop)", "tier": "Venue Partner", "category": "Coworking", "sector": "Coworking", "key_contact": "", "email": "", "website": "https://wework.com", "named_event": "VC Fireside Chat Mixer", "multi_year": "Yes", "ad_placement": "Venue branding", "notes": "NorthLoop Minneapolis; hosted VC mixer"},
    {"year": 2023, "organization": "Fredrikson & Byron P.A.", "tier": "Venue / Session Sponsor", "category": "Legal / Law Firm", "sector": "Professional Services", "key_contact": "", "email": "", "website": "https://fredlaw.com", "named_event": "", "multi_year": "Yes", "ad_placement": "Venue branding", "notes": "Finance & legal session host"},
    {"year": 2023, "organization": "ConnectUP! Institute", "tier": "Community Partner", "category": "Community / DEI", "sector": "Ecosystem", "key_contact": "", "email": "", "website": "", "named_event": "Saint Paul Location", "multi_year": "No", "ad_placement": "Event", "notes": "Saint Paul community location host"},
    {"year": 2023, "organization": "ACER Inc", "tier": "Community Partner", "category": "Community / DEI", "sector": "Ecosystem", "key_contact": "", "email": "", "website": "", "named_event": "Cultural Communities Session", "multi_year": "No", "ad_placement": "Event", "notes": "Cultural communities session host"},
    {"year": 2023, "organization": "Fearless Commerce", "tier": "Community Partner", "category": "Community / Retail", "sector": "Ecosystem", "key_contact": "", "email": "", "website": "", "named_event": "The Watering Event", "multi_year": "No", "ad_placement": "Event", "notes": "The Watering event presenter"},
    {"year": 2023, "organization": "First Independence Bank", "tier": "Community Partner", "category": "Financial Services", "sector": "Banking", "key_contact": "", "email": "", "website": "https://firstindependencebank.com", "named_event": "Business Credit Session", "multi_year": "No", "ad_placement": "Event", "notes": "Business credit session host"},
    {"year": 2023, "organization": "BauHaus Brew Labs", "tier": "Venue Partner", "category": "Food & Beverage", "sector": "Venue", "key_contact": "", "email": "", "website": "https://bauhausbrewlabs.com", "named_event": "Medical Device Networking Happy Hour", "multi_year": "Yes", "ad_placement": "Venue branding", "notes": "Medical Device Networking happy hour venue"},
    {"year": 2023, "organization": "Sociable Ciderworks", "tier": "Venue Partner", "category": "Food & Beverage", "sector": "Venue", "key_contact": "", "email": "", "website": "https://sociablecider.com", "named_event": "OnlyFounders Bags Tournament", "multi_year": "No", "ad_placement": "Venue branding", "notes": "OnlyFounders Bags Tournament venue"},
    {"year": 2023, "organization": "Acme Comedy Company", "tier": "Venue Partner", "category": "Entertainment / Venue", "sector": "Venue", "key_contact": "", "email": "", "website": "https://acmecomedycompany.com", "named_event": "VC Day Happy Hour", "multi_year": "No", "ad_placement": "Venue branding", "notes": "VC Day happy hour venue"},
    {"year": 2023, "organization": "Riverview Theater", "tier": "Venue Partner", "category": "Entertainment / Venue", "sector": "Venue", "key_contact": "", "email": "", "website": "https://riverviewtheater.com", "named_event": "Minnedemo39", "multi_year": "No", "ad_placement": "Venue branding", "notes": "Minnedemo39 venue"},
    {"year": 2023, "organization": "Target Field / MN Twins", "tier": "Venue Partner", "category": "Sports / Entertainment", "sector": "Venue", "key_contact": "", "email": "", "website": "https://mlb.com/twins", "named_event": "BETA Fall Showcase 2023", "multi_year": "No", "ad_placement": "Venue branding", "notes": "BETA Fall Showcase held at Target Field"},
    {"year": 2023, "organization": "McNamara Alumni Center (U of MN)", "tier": "Venue Partner", "category": "Education / Venue", "sector": "Venue", "key_contact": "", "email": "", "website": "https://mcnamaraalumnicenter.com", "named_event": "MN Cup Grand Finale", "multi_year": "No", "ad_placement": "Venue branding", "notes": "MN Cup Grand Finale venue"},
    {"year": 2023, "organization": "Woulfe Alumni Hall / U of St. Thomas", "tier": "Venue Partner", "category": "Education / Venue", "sector": "Venue", "key_contact": "", "email": "", "website": "https://stthomas.edu", "named_event": "gBETA Showcase", "multi_year": "No", "ad_placement": "Venue branding", "notes": "gBETA showcase venue"},
    {"year": 2023, "organization": "IDS Tower / Fulcrum", "tier": "Venue Partner", "category": "Real Estate / Tech", "sector": "Venue", "key_contact": "", "email": "", "website": "", "named_event": "46th Floor Networking Event", "multi_year": "No", "ad_placement": "Venue branding", "notes": "Fulcrum HQ 46th floor networking event"},
    {"year": 2023, "organization": "Northrop / U of Minnesota", "tier": "Venue Partner", "category": "Education / Venue", "sector": "Venue", "key_contact": "", "email": "", "website": "https://northrop.umn.edu", "named_event": "Infrastructure + Future Sessions", "multi_year": "No", "ad_placement": "Venue branding", "notes": "Infrastructure + future sessions 2023"},
    {"year": 2023, "organization": "Hewing Hotel", "tier": "Venue Partner", "category": "Hospitality / Venue", "sector": "Venue", "key_contact": "", "email": "", "website": "https://hewinghotel.com", "named_event": "AR Session", "multi_year": "No", "ad_placement": "Venue branding", "notes": "AR session venue"},
    {"year": 2023, "organization": "MN Cup / University of Minnesota", "tier": "Competition Partner", "category": "Education / Competition", "sector": "Competition / University", "key_contact": "", "email": "hellomncup@umn.edu", "website": "https://mncup.org", "named_event": "Grand Finale ($400K+ prizes)", "multi_year": "Yes", "ad_placement": "All materials", "notes": "Annual pitch competition Grand Finale"},
    # ── 2022 ─────────────────────────────────────────────────────────────────
    {"year": 2022, "organization": "BETA.MN (The Beta Group)", "tier": "Organizer / Operator", "category": "Nonprofit / Community", "sector": "Nonprofit", "key_contact": "Angela Eifert — Executive Director", "email": "Angela.Eifert@BETA.MN", "website": "https://www.beta.mn", "named_event": "", "multi_year": "Yes", "ad_placement": "All materials", "notes": "Primary organizer"},
    {"year": 2022, "organization": "Twin Ignition", "tier": "Presenting Sponsor — Opening Party", "category": "Ecosystem / Community", "sector": "Ecosystem / Community", "key_contact": "", "email": "", "website": "https://twinignition.com", "named_event": "Opening Party", "multi_year": "Yes", "ad_placement": "Named event + website", "notes": "Opening party co-host; Startup Garage + Keg House venues"},
    {"year": 2022, "organization": "Workbox", "tier": "Venue Sponsor", "category": "Coworking / Real Estate", "sector": "Coworking", "key_contact": "", "email": "", "website": "https://workbox.com", "named_event": "", "multi_year": "Yes", "ad_placement": "Venue branding", "notes": "Event venue"},
    {"year": 2022, "organization": "FINNOVATION Lab", "tier": "Venue Sponsor", "category": "Fintech", "sector": "Fintech", "key_contact": "", "email": "", "website": "https://finnovationlab.com", "named_event": "", "multi_year": "Yes", "ad_placement": "Venue branding", "notes": "Training Lab venue"},
    {"year": 2022, "organization": "MN Cup / University of Minnesota", "tier": "Competition Partner", "category": "Education / Competition", "sector": "Competition", "key_contact": "", "email": "hellomncup@umn.edu", "website": "https://mncup.org", "named_event": "Grand Finale", "multi_year": "Yes", "ad_placement": "All materials", "notes": "Annual pitch competition"},
    # ── 2021 ─────────────────────────────────────────────────────────────────
    {"year": 2021, "organization": "BETA.MN (The Beta Group)", "tier": "Organizer / Operator", "category": "Nonprofit / Community", "sector": "Nonprofit", "key_contact": "Angela Eifert — Executive Director", "email": "Angela.Eifert@BETA.MN", "website": "https://www.beta.mn", "named_event": "", "multi_year": "Yes", "ad_placement": "All materials", "notes": "Primary organizer; virtual/hybrid event"},
    {"year": 2021, "organization": "MN Cup / University of Minnesota", "tier": "Competition Partner", "category": "Education / Competition", "sector": "Competition", "key_contact": "", "email": "hellomncup@umn.edu", "website": "https://mncup.org", "named_event": "Grand Finale", "multi_year": "Yes", "ad_placement": "All materials", "notes": "Annual pitch competition"},
    # ── 2020 ─────────────────────────────────────────────────────────────────
    {"year": 2020, "organization": "BETA.MN (The Beta Group)", "tier": "Organizer / Operator", "category": "Nonprofit / Community", "sector": "Nonprofit", "key_contact": "Angela Eifert — Executive Director", "email": "Angela.Eifert@BETA.MN", "website": "https://www.beta.mn", "named_event": "", "multi_year": "Yes", "ad_placement": "All materials", "notes": "Primary organizer; COVID — virtual event"},
    {"year": 2020, "organization": "MN Cup / University of Minnesota", "tier": "Competition Partner", "category": "Education / Competition", "sector": "Competition", "key_contact": "", "email": "hellomncup@umn.edu", "website": "https://mncup.org", "named_event": "Grand Finale", "multi_year": "Yes", "ad_placement": "All materials", "notes": "Annual pitch competition"},
    # ── 2019 ─────────────────────────────────────────────────────────────────
    {"year": 2019, "organization": "BETA.MN (The Beta Group)", "tier": "Organizer / Operator", "category": "Nonprofit / Community", "sector": "Nonprofit", "key_contact": "Angela Eifert — Executive Director", "email": "Angela.Eifert@BETA.MN", "website": "https://www.beta.mn", "named_event": "", "multi_year": "Yes", "ad_placement": "All materials", "notes": "Primary organizer"},
    {"year": 2019, "organization": "Target Corporation", "tier": "Title Sponsor", "category": "Retail / Fortune 500", "sector": "Retail", "key_contact": "Gene Han — VP New Ventures & Accelerators", "email": "", "website": "https://target.com", "named_event": "Kickoff Event + Retail Track at Target Plaza Commons", "multi_year": "No", "ad_placement": "All materials + named venue", "notes": "Hosted Kickoff Event at Target HQ; community investment + innovation pipeline"},
    {"year": 2019, "organization": "Cargill", "tier": "Title Sponsor", "category": "Agriculture / Food / Fortune 500", "sector": "Agriculture", "key_contact": "Lawrence Wang — Digitalization & Analytics Strategy Lead", "email": "", "website": "https://cargill.com", "named_event": "Techstars Farm to Fork Demo Day (Oct 15)", "multi_year": "Yes", "ad_placement": "All materials + named event", "notes": "Co-sponsored Techstars Farm to Fork; Wayzata MN HQ"},
    {"year": 2019, "organization": "3M", "tier": "Title Sponsor", "category": "Manufacturing / Science / Fortune 500", "sector": "Manufacturing", "key_contact": "Ben Wright — Director of 3M Ventures", "email": "", "website": "https://3m.com", "named_event": "BETA Backers Program", "multi_year": "Yes", "ad_placement": "All materials + named program", "notes": "3M Ventures deal flow access; BETA Backers Program"},
    {"year": 2019, "organization": "AWS (Amazon Web Services)", "tier": "Corporate Sponsor", "category": "Technology / Cloud", "sector": "Technology", "key_contact": "", "email": "", "website": "https://aws.amazon.com", "named_event": "", "multi_year": "No", "ad_placement": "Website + materials", "notes": "Corporate sponsor 2019"},
    {"year": 2019, "organization": "Microsoft", "tier": "Corporate Sponsor", "category": "Technology / Software", "sector": "Technology", "key_contact": "", "email": "", "website": "https://microsoft.com", "named_event": "", "multi_year": "No", "ad_placement": "Website + materials", "notes": "Corporate sponsor 2019"},
    {"year": 2019, "organization": "Ecolab", "tier": "Event Co-Sponsor", "category": "Industrial / Sustainability", "sector": "Industrial", "key_contact": "", "email": "", "website": "https://ecolab.com", "named_event": "", "multi_year": "No", "ad_placement": "Event", "notes": "Event co-sponsor 2019"},
    {"year": 2019, "organization": "GreaterMSP", "tier": "Presenting / Promotional Partner", "category": "Economic Development", "sector": "Economic Dev", "key_contact": "", "email": "", "website": "https://greatermsp.org", "named_event": "", "multi_year": "No", "ad_placement": "Promotional", "notes": "Regional economic development partner"},
    {"year": 2019, "organization": "Make It. MSP", "tier": "Fly-In Program Sponsor", "category": "Economic Development", "sector": "Economic Dev", "key_contact": "", "email": "", "website": "", "named_event": "Fly-In Program", "multi_year": "Yes", "ad_placement": "Event", "notes": "Flew in non-Minnesota job seekers to experience TCSW"},
    {"year": 2019, "organization": "Branch Messenger", "tier": "Fly-In Program Sponsor", "category": "Technology / Startup", "sector": "Technology", "key_contact": "", "email": "", "website": "", "named_event": "Fly-In Program", "multi_year": "Yes", "ad_placement": "Event", "notes": "Fly-in program co-sponsor"},
    {"year": 2019, "organization": "Techstars", "tier": "Event Partner", "category": "Startup Accelerator", "sector": "Accelerator", "key_contact": "", "email": "", "website": "https://techstars.com", "named_event": "Farm to Fork Demo Day", "multi_year": "No", "ad_placement": "Event + website", "notes": "Techstars Farm to Fork Demo Day"},
    {"year": 2019, "organization": "MN Cup / University of Minnesota", "tier": "Competition Partner", "category": "Education / Competition", "sector": "Competition", "key_contact": "", "email": "hellomncup@umn.edu", "website": "https://mncup.org", "named_event": "Grand Finale", "multi_year": "Yes", "ad_placement": "All materials", "notes": "Annual pitch competition"},
    {"year": 2019, "organization": "Launch Minnesota", "tier": "Government / Ecosystem Partner", "category": "Government", "sector": "Government", "key_contact": "", "email": "", "website": "https://launchmn.com", "named_event": "", "multi_year": "Yes", "ad_placement": "Website", "notes": "State of MN startup initiative"},
    # ── 2018 ─────────────────────────────────────────────────────────────────
    {"year": 2018, "organization": "BETA.MN (The Beta Group)", "tier": "Organizer / Operator", "category": "Nonprofit / Community", "sector": "Nonprofit", "key_contact": "Angela Eifert — Executive Director", "email": "Angela.Eifert@BETA.MN", "website": "https://www.beta.mn", "named_event": "", "multi_year": "Yes", "ad_placement": "All materials", "notes": "Primary organizer"},
    {"year": 2018, "organization": "Lyft", "tier": "Transport Sponsor", "category": "Technology / Transportation", "sector": "Technology", "key_contact": "", "email": "", "website": "https://lyft.com", "named_event": "Free Lyft Bus Shuttle", "multi_year": "No", "ad_placement": "Event + app", "notes": "Free Lyft bus shuttle during TCSW 2018"},
    {"year": 2018, "organization": "Mayo Clinic", "tier": "MANOVA Summit Sponsor", "category": "Healthcare", "sector": "Healthcare", "key_contact": "", "email": "", "website": "https://mayoclinic.org", "named_event": "MANOVA Summit", "multi_year": "No", "ad_placement": "Event", "notes": "MANOVA healthcare summit sponsor"},
    {"year": 2018, "organization": "Walmart", "tier": "MANOVA Summit Sponsor", "category": "Retail / Fortune 500", "sector": "Retail", "key_contact": "", "email": "", "website": "https://walmart.com", "named_event": "MANOVA Summit", "multi_year": "No", "ad_placement": "Event", "notes": "MANOVA summit sponsor"},
    {"year": 2018, "organization": "Cargill", "tier": "Co-Sponsor", "category": "Agriculture / Food", "sector": "Agriculture", "key_contact": "", "email": "", "website": "https://cargill.com", "named_event": "", "multi_year": "Yes", "ad_placement": "Event", "notes": "Co-sponsor 2018"},
    {"year": 2018, "organization": "3M", "tier": "Sponsor", "category": "Manufacturing / Science", "sector": "Manufacturing", "key_contact": "", "email": "", "website": "https://3m.com", "named_event": "", "multi_year": "Yes", "ad_placement": "Event", "notes": "Sponsor 2018"},
    {"year": 2018, "organization": "MN Cup / University of Minnesota", "tier": "Competition Partner", "category": "Education / Competition", "sector": "Competition", "key_contact": "", "email": "hellomncup@umn.edu", "website": "https://mncup.org", "named_event": "Grand Finale", "multi_year": "Yes", "ad_placement": "All materials", "notes": "Annual pitch competition"},
    {"year": 2018, "organization": "BauHaus Brew Labs", "tier": "Venue Partner", "category": "Food & Beverage", "sector": "Venue", "key_contact": "", "email": "", "website": "https://bauhausbrewlabs.com", "named_event": "", "multi_year": "Yes", "ad_placement": "Venue branding", "notes": "Community venue partner"},
    # ── 2017 ─────────────────────────────────────────────────────────────────
    {"year": 2017, "organization": "BETA.MN (The Beta Group)", "tier": "Organizer / Operator", "category": "Nonprofit / Community", "sector": "Nonprofit", "key_contact": "", "email": "", "website": "https://www.beta.mn", "named_event": "", "multi_year": "Yes", "ad_placement": "All materials", "notes": "Primary organizer"},
    {"year": 2017, "organization": "Allianz Life", "tier": "Major Sponsor — Innovation Day Host", "category": "Insurance / Financial Services", "sector": "Financial Services", "key_contact": "", "email": "", "website": "https://allianzlife.com", "named_event": "Innovation Day", "multi_year": "No", "ad_placement": "Named event + venue", "notes": "Hosted Innovation Day 2017"},
    {"year": 2017, "organization": "Avisen Legal P.A.", "tier": "Legal Sponsor", "category": "Legal / Law Firm", "sector": "Professional Services", "key_contact": "", "email": "", "website": "", "named_event": "", "multi_year": "No", "ad_placement": "Website + materials", "notes": "Legal sponsor 2017"},
    {"year": 2017, "organization": "Bauhaus Brewery", "tier": "Venue Partner", "category": "Food & Beverage", "sector": "Venue", "key_contact": "", "email": "", "website": "https://bauhausbrewlabs.com", "named_event": "", "multi_year": "Yes", "ad_placement": "Venue branding", "notes": "Community venue partner"},
    {"year": 2017, "organization": "CoCo (Coworking Community)", "tier": "Venue Partner — Co-Working", "category": "Coworking", "sector": "Coworking", "key_contact": "", "email": "", "website": "", "named_event": "", "multi_year": "No", "ad_placement": "Venue branding", "notes": "Co-working venue 2017"},
    {"year": 2017, "organization": "Fredrikson & Byron P.A.", "tier": "Legal Sponsor", "category": "Legal / Law Firm", "sector": "Professional Services", "key_contact": "", "email": "", "website": "https://fredlaw.com", "named_event": "", "multi_year": "Yes", "ad_placement": "Website + materials", "notes": "Legal sponsor"},
    {"year": 2017, "organization": "Industrious", "tier": "Venue Partner — Co-Working", "category": "Coworking", "sector": "Coworking", "key_contact": "", "email": "", "website": "https://industriousoffice.com", "named_event": "", "multi_year": "No", "ad_placement": "Venue branding", "notes": "Co-working venue 2017"},
    {"year": 2017, "organization": "Mall of America", "tier": "Venue Partner", "category": "Retail / Events", "sector": "Retail", "key_contact": "", "email": "", "website": "https://mallofamerica.com", "named_event": "", "multi_year": "No", "ad_placement": "Venue branding", "notes": "Event venue 2017"},
    {"year": 2017, "organization": "WeWork", "tier": "Venue Partner — Co-Working", "category": "Coworking", "sector": "Coworking", "key_contact": "", "email": "", "website": "https://wework.com", "named_event": "", "multi_year": "Yes", "ad_placement": "Venue branding", "notes": "Co-working venue 2017 and 2023"},
    {"year": 2017, "organization": "Make It. MSP", "tier": "Fly-In Program Partner", "category": "Economic Development", "sector": "Economic Dev", "key_contact": "", "email": "", "website": "", "named_event": "Fly-In Program", "multi_year": "Yes", "ad_placement": "Event", "notes": "Fly-in talent attraction program"},
    # ── 2014 (Founding) ──────────────────────────────────────────────────────
    {"year": 2014, "organization": "BETA.MN (The Beta Group)", "tier": "Founding Organizer", "category": "Nonprofit / Community", "sector": "Nonprofit", "key_contact": "", "email": "", "website": "https://www.beta.mn", "named_event": "", "multi_year": "Yes", "ad_placement": "All materials", "notes": "Founded TCSW in 2014"},
    {"year": 2014, "organization": "Fueled Collective", "tier": "Founding Partner", "category": "Coworking", "sector": "Coworking", "key_contact": "", "email": "", "website": "", "named_event": "", "multi_year": "No", "ad_placement": "Founding materials", "notes": "Founding coworking partner"},
    {"year": 2014, "organization": "Minnestar", "tier": "Founding Partner", "category": "Ecosystem / Community", "sector": "Ecosystem", "key_contact": "", "email": "", "website": "", "named_event": "", "multi_year": "No", "ad_placement": "Founding materials", "notes": "Founding community ecosystem partner"},
    {"year": 2014, "organization": "TECHdotMN", "tier": "Founding Media Partner", "category": "Media / Technology", "sector": "Media", "key_contact": "", "email": "", "website": "https://techdotmn.com", "named_event": "", "multi_year": "Yes", "ad_placement": "Media coverage", "notes": "Founding community media partner; annual supporter"},
    {"year": 2014, "organization": "MN Cup / University of Minnesota", "tier": "Competition Partner", "category": "Education / Competition", "sector": "Competition", "key_contact": "", "email": "hellomncup@umn.edu", "website": "https://mncup.org", "named_event": "Grand Finale", "multi_year": "Yes", "ad_placement": "All materials", "notes": "Annual pitch competition; present since founding"},
]

FIELDS = ["year", "organization", "tier", "category", "sector", "key_contact",
          "email", "website", "named_event", "multi_year", "ad_placement", "notes"]


# ══════════════════════════════════════════════════════════════════════════════
# BUILD XLSX
# ══════════════════════════════════════════════════════════════════════════════

def build_xlsx():
    import openpyxl
    from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                                  GradientFill)
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()

    # ── Colors ──────────────────────────────────────────────────────────────
    DARK_BLUE  = "1B2A4A"
    MID_BLUE   = "2E4A7A"
    LIGHT_BLUE = "D6E4F0"
    GOLD       = "C9A84C"
    WHITE      = "FFFFFF"
    GREY_ROW   = "F4F6F9"
    BORDER_COL = "C0CCDA"

    thin = Border(
        left=Side(style='thin', color=BORDER_COL),
        right=Side(style='thin', color=BORDER_COL),
        top=Side(style='thin', color=BORDER_COL),
        bottom=Side(style='thin', color=BORDER_COL)
    )

    def hdr_cell(ws, row, col, val, bg=DARK_BLUE, fg=WHITE, bold=True, size=11):
        c = ws.cell(row=row, column=col, value=val)
        c.font = Font(bold=bold, color=fg, size=size, name="Calibri")
        c.fill = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = thin
        return c

    def data_cell(ws, row, col, val, bg=WHITE, bold=False, wrap=True):
        c = ws.cell(row=row, column=col, value=val)
        c.font = Font(bold=bold, color="1B2A4A", size=10, name="Calibri")
        c.fill = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(vertical="center", wrap_text=wrap)
        c.border = thin
        return c

    # ════════════════════════════════════════════════════════════════════════
    # SHEET 1 — Master All-Years Table
    # ════════════════════════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "Master — All Years"

    # Title banner
    ws1.merge_cells("A1:L1")
    title = ws1["A1"]
    title.value = "TWIN CITIES STARTUP WEEK — COMPLETE SPONSOR & PARTNER MASTER LIST"
    title.font = Font(bold=True, color=WHITE, size=14, name="Calibri")
    title.fill = PatternFill("solid", fgColor=DARK_BLUE)
    title.alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 30

    ws1.merge_cells("A2:L2")
    sub = ws1["A2"]
    sub.value = f"All confirmed sponsors, partners, venues, and co-hosts from 2014–2026 | Generated {TODAY} | Sources: tcstartupweek.com, beta.mn, tcsw2023.sched.com, tcbmag.com, start-midwest.com"
    sub.font = Font(italic=True, color="555555", size=9, name="Calibri")
    sub.fill = PatternFill("solid", fgColor="EEF2F7")
    sub.alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[2].height = 18

    # Column headers
    HEADERS_ROW = ["Year", "Organization", "Tier / Role", "Category", "Sector",
                   "Key Contact", "Email", "Website", "Named Event / Activation",
                   "Multi-Year?", "Ad Placement", "Notes"]
    col_widths   = [7, 32, 28, 22, 18, 30, 26, 30, 28, 10, 18, 40]

    for c, (h, w) in enumerate(zip(HEADERS_ROW, col_widths), 1):
        hdr_cell(ws1, 3, c, h, bg=DARK_BLUE, size=10)
        ws1.column_dimensions[get_column_letter(c)].width = w
    ws1.row_dimensions[3].height = 22

    # Data rows — color by year band
    YEAR_COLORS = {
        2026: "EBF5FB", 2025: "E8F8F5", 2024: "FEF9E7",
        2023: "FDEDEC", 2022: "F4ECF7", 2021: "EBF5FB",
        2020: "E8F8F5", 2019: "FEF9E7", 2018: "FDEDEC",
        2017: "F4ECF7", 2014: "FDFEFE",
    }

    sorted_data = sorted(MASTER, key=lambda x: (-x["year"], x["organization"]))
    prev_year   = None

    for r_idx, row in enumerate(sorted_data, 4):
        yr   = row["year"]
        bg   = YEAR_COLORS.get(yr, WHITE)

        # Year divider row
        if yr != prev_year:
            ws1.merge_cells(f"A{r_idx}:L{r_idx}")
            yc = ws1[f"A{r_idx}"]
            count = sum(1 for d in MASTER if d["year"] == yr)
            yc.value = f"  ▶  {yr}  —  {count} Sponsors / Partners"
            yc.font  = Font(bold=True, color=WHITE, size=10, name="Calibri")
            yc.fill  = PatternFill("solid", fgColor=MID_BLUE)
            yc.alignment = Alignment(vertical="center")
            ws1.row_dimensions[r_idx].height = 18
            r_idx += 1
            prev_year = yr

        for c_idx, field in enumerate(FIELDS, 1):
            val = row.get(field, "")
            bold_col = c_idx == 2  # bold org name
            data_cell(ws1, r_idx, c_idx, val, bg=bg, bold=bold_col)
        ws1.row_dimensions[r_idx].height = 28

    ws1.freeze_panes = "A4"

    # ════════════════════════════════════════════════════════════════════════
    # SHEET 2 — By Year Summary
    # ════════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("By Year Summary")
    ws2.merge_cells("A1:F1")
    t2 = ws2["A1"]
    t2.value = "TCSW SPONSORS — YEAR-BY-YEAR SUMMARY"
    t2.font  = Font(bold=True, color=WHITE, size=13, name="Calibri")
    t2.fill  = PatternFill("solid", fgColor=DARK_BLUE)
    t2.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 28
    for c, w in zip(range(1, 7), [30, 28, 22, 30, 24, 42]):
        ws2.column_dimensions[get_column_letter(c)].width = w

    current_row = 2
    for yr in sorted(set(d["year"] for d in MASTER), reverse=True):
        year_rows = sorted([d for d in MASTER if d["year"] == yr], key=lambda x: x["organization"])

        # Year header
        ws2.merge_cells(f"A{current_row}:F{current_row}")
        yh = ws2[f"A{current_row}"]
        yh.value = f"  {yr}  —  {len(year_rows)} Sponsors / Partners"
        yh.font  = Font(bold=True, color=WHITE, size=11, name="Calibri")
        yh.fill  = PatternFill("solid", fgColor=MID_BLUE)
        yh.alignment = Alignment(vertical="center")
        ws2.row_dimensions[current_row].height = 20
        current_row += 1

        # Sub-headers
        for c, h in enumerate(["Organization", "Tier / Role", "Category", "Key Contact", "Named Event", "Notes"], 1):
            hdr_cell(ws2, current_row, c, h, bg=LIGHT_BLUE, fg=DARK_BLUE, size=9)
        ws2.row_dimensions[current_row].height = 18
        current_row += 1

        for i, row in enumerate(year_rows):
            bg = WHITE if i % 2 == 0 else GREY_ROW
            for c, field in enumerate(["organization", "tier", "category", "key_contact", "named_event", "notes"], 1):
                data_cell(ws2, current_row, c, row.get(field, ""), bg=bg)
            ws2.row_dimensions[current_row].height = 22
            current_row += 1

        current_row += 1  # gap

    # ════════════════════════════════════════════════════════════════════════
    # SHEET 3 — Fortune 500 Spotlight
    # ════════════════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("Fortune 500 Spotlight")
    f500 = [d for d in MASTER if any(x in d["tier"].lower() for x in ["title", "corporate", "premier"]) and d["year"] in [2018, 2019]]
    ws3.merge_cells("A1:E1")
    t3 = ws3["A1"]
    t3.value = "TCSW FORTUNE 500 & MAJOR CORPORATE SPONSORS"
    t3.font  = Font(bold=True, color=WHITE, size=13, name="Calibri")
    t3.fill  = PatternFill("solid", fgColor=DARK_BLUE)
    t3.alignment = Alignment(horizontal="center", vertical="center")
    ws3.row_dimensions[1].height = 28
    for c, (h, w) in enumerate(zip(["Year", "Organization", "Tier", "Key Contact", "Notes / Activation"], [7, 28, 22, 32, 50]), 1):
        hdr_cell(ws3, 2, c, h, bg=MID_BLUE, size=10)
        ws3.column_dimensions[get_column_letter(c)].width = w
    ws3.row_dimensions[2].height = 20
    for i, row in enumerate(f500, 3):
        bg = WHITE if i % 2 == 0 else GREY_ROW
        for c, field in enumerate(["year", "organization", "tier", "key_contact", "notes"], 1):
            data_cell(ws3, i, c, row.get(field, ""), bg=bg)
        ws3.row_dimensions[i].height = 24

    # ════════════════════════════════════════════════════════════════════════
    # SHEET 4 — Multi-Year Sponsors
    # ════════════════════════════════════════════════════════════════════════
    ws4 = wb.create_sheet("Multi-Year Sponsors")
    multi = {}
    for row in MASTER:
        org = row["organization"]
        if org not in multi:
            multi[org] = {"years": set(), "tiers": set(), "sector": row["sector"],
                          "website": row["website"], "key_contact": row["key_contact"],
                          "email": row["email"]}
        multi[org]["years"].add(row["year"])
        multi[org]["tiers"].add(row["tier"])
    multi_sorted = sorted([(k, v) for k, v in multi.items() if len(v["years"]) > 1],
                          key=lambda x: -len(x[1]["years"]))

    ws4.merge_cells("A1:G1")
    t4 = ws4["A1"]
    t4.value = f"TCSW MULTI-YEAR SPONSORS & PARTNERS ({len(multi_sorted)} organizations)"
    t4.font  = Font(bold=True, color=WHITE, size=13, name="Calibri")
    t4.fill  = PatternFill("solid", fgColor=DARK_BLUE)
    t4.alignment = Alignment(horizontal="center", vertical="center")
    ws4.row_dimensions[1].height = 28
    for c, (h, w) in enumerate(zip(["Organization", "Years Active", "# Years", "Sector", "Highest Role", "Key Contact", "Website"], [30, 18, 8, 20, 28, 30, 30]), 1):
        hdr_cell(ws4, 2, c, h, bg=MID_BLUE, size=10)
        ws4.column_dimensions[get_column_letter(c)].width = w
    ws4.row_dimensions[2].height = 20
    for i, (org, d) in enumerate(multi_sorted, 3):
        bg = WHITE if i % 2 == 0 else GREY_ROW
        years_str = ", ".join(str(y) for y in sorted(d["years"], reverse=True))
        tier_str  = " / ".join(sorted(d["tiers"]))[:50]
        for c, val in enumerate([org, years_str, len(d["years"]), d["sector"], tier_str, d["key_contact"], d["website"]], 1):
            data_cell(ws4, i, c, val, bg=bg, bold=(c == 1))
        ws4.row_dimensions[i].height = 22

    wb.save(OUT / "TCSW_Sponsors_Master.xlsx")
    print(f"  ✓ TCSW_Sponsors_Master.xlsx — {len(MASTER)} rows across 4 sheets")


# ══════════════════════════════════════════════════════════════════════════════
# BUILD PDF
# ══════════════════════════════════════════════════════════════════════════════

def build_pdf():
    from reportlab.lib.pagesizes import letter, landscape
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle, Paragraph,
                                     Spacer, PageBreak, HRFlowable)
    from reportlab.lib.enums import TA_CENTER, TA_LEFT

    doc = SimpleDocTemplate(
        str(OUT / "TCSW_Sponsors_Master.pdf"),
        pagesize=landscape(letter),
        rightMargin=0.4*inch, leftMargin=0.4*inch,
        topMargin=0.4*inch, bottomMargin=0.4*inch,
        title="TCSW Sponsors Master List",
        author="TCSW Research",
    )

    DARK  = colors.HexColor("#1B2A4A")
    MID   = colors.HexColor("#2E4A7A")
    GOLD  = colors.HexColor("#C9A84C")
    LBLUE = colors.HexColor("#D6E4F0")
    GREY  = colors.HexColor("#F4F6F9")
    WHITE = colors.white

    YEAR_BG = {
        2026: colors.HexColor("#EBF5FB"),
        2025: colors.HexColor("#E8F8F5"),
        2024: colors.HexColor("#FEF9E7"),
        2023: colors.HexColor("#FDEDEC"),
        2022: colors.HexColor("#F4ECF7"),
        2021: colors.HexColor("#EBF5FB"),
        2020: colors.HexColor("#E8F8F5"),
        2019: colors.HexColor("#FEF9E7"),
        2018: colors.HexColor("#FDEDEC"),
        2017: colors.HexColor("#F4ECF7"),
        2014: colors.HexColor("#FDFEFE"),
    }

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("Title", fontSize=18, textColor=WHITE,
                                  alignment=TA_CENTER, fontName="Helvetica-Bold",
                                  spaceAfter=4)
    sub_style   = ParagraphStyle("Sub", fontSize=8, textColor=colors.HexColor("#555555"),
                                  alignment=TA_CENTER, fontName="Helvetica")
    h1_style    = ParagraphStyle("H1", fontSize=13, textColor=DARK,
                                  fontName="Helvetica-Bold", spaceBefore=12, spaceAfter=4)
    h2_style    = ParagraphStyle("H2", fontSize=10, textColor=MID,
                                  fontName="Helvetica-Bold", spaceBefore=6, spaceAfter=2)
    body_style  = ParagraphStyle("Body", fontSize=9, textColor=DARK,
                                  fontName="Helvetica", leading=13)
    small_style = ParagraphStyle("Small", fontSize=7.5, textColor=colors.HexColor("#444"),
                                  fontName="Helvetica", leading=10)

    story = []

    # ── Cover ──────────────────────────────────────────────────────────────
    def cover_bg(canvas, doc):
        canvas.saveState()
        canvas.setFillColor(DARK)
        canvas.rect(0, 0, landscape(letter)[0], landscape(letter)[1], fill=1, stroke=0)
        canvas.setFillColor(GOLD)
        canvas.rect(0, 3.2*inch, landscape(letter)[0], 0.06*inch, fill=1, stroke=0)
        canvas.restoreState()

    cover_data = [[Paragraph("TWIN CITIES STARTUP WEEK", title_style)],
                  [Paragraph("COMPLETE SPONSOR &amp; PARTNER MASTER LIST", title_style)],
                  [Paragraph(f"2014 – 2026  ·  {len(MASTER)} Entries  ·  Generated {TODAY}", sub_style)]]
    cover_tbl = Table(cover_data, colWidths=[10*inch])
    cover_tbl.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,-1), DARK),
        ("ALIGN",      (0,0), (-1,-1), "CENTER"),
        ("TOPPADDING", (0,0), (-1,-1), 14),
        ("BOTTOMPADDING", (0,0), (-1,-1), 14),
    ]))
    story.append(Spacer(1, 2.2*inch))
    story.append(cover_tbl)
    story.append(PageBreak())

    # ── Section 1: Stats ────────────────────────────────────────────────────
    story.append(Paragraph("SECTION 1 — KEY STATISTICS", h1_style))
    story.append(HRFlowable(width="100%", thickness=1, color=GOLD, spaceAfter=8))

    years_present = sorted(set(d["year"] for d in MASTER), reverse=True)
    unique_orgs   = len(set(d["organization"] for d in MASTER))
    multi_count   = sum(1 for org in set(d["organization"] for d in MASTER)
                        if len([d for d in MASTER if d["organization"] == org]) > 1)
    named_events  = len([d for d in MASTER if d["named_event"]])
    stats = [
        ["Total Years Covered", str(len(years_present)),
         "Unique Organizations", str(unique_orgs)],
        ["Total Sponsor Records", str(len(MASTER)),
         "Multi-Year Partners", str(multi_count)],
        ["Named Events / Activations", str(named_events),
         "Fortune 500 Sponsors", "5 (Target, Cargill, 3M, Walmart, AWS/Microsoft)"],
    ]
    stats_tbl = Table(stats, colWidths=[2.2*inch, 1.2*inch, 2.2*inch, 4.4*inch])
    stats_tbl.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,-1), GREY),
        ("TEXTCOLOR",  (0,0), (-1,-1), DARK),
        ("FONTNAME",   (0,0), (0,-1), "Helvetica-Bold"),
        ("FONTNAME",   (2,0), (2,-1), "Helvetica-Bold"),
        ("FONTSIZE",   (0,0), (-1,-1), 9),
        ("GRID",       (0,0), (-1,-1), 0.5, colors.HexColor("#C0CCDA")),
        ("ROWBACKGROUNDS", (0,0), (-1,-1), [WHITE, GREY]),
        ("PADDING",    (0,0), (-1,-1), 6),
    ]))
    story.append(stats_tbl)
    story.append(PageBreak())

    # ── Section 2: Complete Master Table ───────────────────────────────────
    story.append(Paragraph("SECTION 2 — COMPLETE MASTER TABLE (ALL YEARS)", h1_style))
    story.append(HRFlowable(width="100%", thickness=1, color=GOLD, spaceAfter=6))

    col_labels = ["Year", "Organization", "Tier / Role", "Category",
                  "Key Contact", "Website", "Named Event", "Multi-Yr?", "Notes"]
    col_fields = ["year", "organization", "tier", "category",
                  "key_contact", "website", "named_event", "multi_year", "notes"]
    col_widths_pt = [0.45*inch, 1.8*inch, 1.7*inch, 1.3*inch,
                     1.6*inch, 1.5*inch, 1.4*inch, 0.55*inch, 1.8*inch]

    def make_cell(text, bold=False, size=8):
        st = ParagraphStyle("c", fontSize=size, fontName="Helvetica-Bold" if bold else "Helvetica",
                             textColor=DARK, leading=10)
        return Paragraph(str(text or ""), st)

    table_data = [[make_cell(h, bold=True) for h in col_labels]]
    sorted_data = sorted(MASTER, key=lambda x: (-x["year"], x["organization"]))
    prev_yr = None

    cmd = [
        ("BACKGROUND", (0,0), (-1,0), MID),
        ("TEXTCOLOR",  (0,0), (-1,0), WHITE),
        ("GRID",       (0,0), (-1,-1), 0.4, colors.HexColor("#C0CCDA")),
        ("FONTSIZE",   (0,0), (-1,-1), 8),
        ("VALIGN",     (0,0), (-1,-1), "MIDDLE"),
        ("TOPPADDING", (0,0), (-1,-1), 3),
        ("BOTTOMPADDING", (0,0), (-1,-1), 3),
    ]

    for r_idx, row in enumerate(sorted_data, 1):
        yr = row["year"]
        if yr != prev_yr:
            year_count = sum(1 for d in MASTER if d["year"] == yr)
            yr_label = [make_cell(f"▶  {yr}  —  {year_count} entries", bold=True, size=8)]
            yr_label += [make_cell("") for _ in range(len(col_labels)-1)]
            table_data.append(yr_label)
            cmd.append(("BACKGROUND", (0, len(table_data)-1), (-1, len(table_data)-1), LBLUE))
            cmd.append(("SPAN", (0, len(table_data)-1), (-1, len(table_data)-1)))
            prev_yr = yr

        cells = [make_cell(row.get(f, ""), bold=(f == "organization")) for f in col_fields]
        bg_c = YEAR_BG.get(yr, WHITE)
        table_data.append(cells)
        if len(table_data) % 2 == 0:
            cmd.append(("BACKGROUND", (0, len(table_data)-1), (-1, len(table_data)-1), bg_c))

    tbl = Table(table_data, colWidths=col_widths_pt, repeatRows=1)
    tbl.setStyle(TableStyle(cmd))
    story.append(tbl)
    story.append(PageBreak())

    # ── Section 3: Fortune 500 Spotlight ───────────────────────────────────
    story.append(Paragraph("SECTION 3 — FORTUNE 500 SPONSOR SPOTLIGHT (2019)", h1_style))
    story.append(HRFlowable(width="100%", thickness=1, color=GOLD, spaceAfter=8))
    f500_notes = [
        ("Target Corporation", "Title Sponsor 2019", "Gene Han — VP New Ventures & Accelerators",
         "Hosted Kickoff Event + Retail Track at Target Plaza Commons (Minneapolis HQ). \"Startups and corporations can work together in a way that is mutually beneficial.\""),
        ("Cargill", "Title Sponsor 2019 (co-sponsor 2018)", "Lawrence Wang — Digitalization & Analytics Strategy Lead",
         "Co-sponsored Techstars Farm to Fork; Demo Day October 15. \"The pace of innovation is so fast that close-knit relationships with early stage founders is required.\""),
        ("3M", "Title Sponsor 2019 (sponsor 2018)", "Ben Wright — Director of 3M Ventures",
         "BETA Backers Program (investors meet early-stage MN startups). \"Twin Cities Startup Week is a great opportunity to collaborate with the local entrepreneurial community.\""),
        ("AWS", "Corporate Sponsor 2019", "—",
         "Cloud infrastructure sponsor; national presence at startup events."),
        ("Microsoft", "Corporate Sponsor 2019", "—",
         "Corporate sponsor; Azure/startup programs alignment."),
        ("Walmart", "MANOVA Summit Sponsor 2018", "—",
         "MANOVA healthcare/retail innovation summit sponsor."),
    ]
    for org, tier, contact, note in f500_notes:
        story.append(Paragraph(f"<b>{org}</b> — {tier}", h2_style))
        story.append(Paragraph(f"<b>Contact:</b> {contact}", small_style))
        story.append(Paragraph(note, small_style))
        story.append(Spacer(1, 6))

    story.append(PageBreak())

    # ── Section 4: Multi-Year Partners ─────────────────────────────────────
    story.append(Paragraph("SECTION 4 — MULTI-YEAR PARTNERS", h1_style))
    story.append(HRFlowable(width="100%", thickness=1, color=GOLD, spaceAfter=6))

    multi_orgs = {}
    for row in MASTER:
        org = row["organization"]
        if org not in multi_orgs:
            multi_orgs[org] = {"years": set(), "sector": row["sector"], "website": row["website"]}
        multi_orgs[org]["years"].add(row["year"])
    multi_list = sorted([(k, v) for k, v in multi_orgs.items() if len(v["years"]) > 1],
                        key=lambda x: -len(x[1]["years"]))

    m_data  = [[make_cell(h, bold=True) for h in ["Organization", "Years Active", "# Years", "Sector", "Website"]]]
    m_cmd   = [("BACKGROUND", (0,0), (-1,0), MID), ("TEXTCOLOR", (0,0), (-1,0), WHITE),
               ("GRID", (0,0), (-1,-1), 0.4, colors.HexColor("#C0CCDA")),
               ("FONTSIZE", (0,0), (-1,-1), 8), ("VALIGN", (0,0), (-1,-1), "MIDDLE")]
    for i, (org, d) in enumerate(multi_list):
        yr_str = ", ".join(str(y) for y in sorted(d["years"], reverse=True))
        m_data.append([make_cell(org, bold=True), make_cell(yr_str),
                        make_cell(str(len(d["years"]))), make_cell(d["sector"]),
                        make_cell(d["website"])])
        if i % 2 == 0:
            m_cmd.append(("BACKGROUND", (0, i+1), (-1, i+1), GREY))

    m_tbl = Table(m_data, colWidths=[2.3*inch, 2.5*inch, 0.6*inch, 1.5*inch, 2.2*inch])
    m_tbl.setStyle(TableStyle(m_cmd))
    story.append(m_tbl)

    doc.build(story)
    print(f"  ✓ TCSW_Sponsors_Master.pdf — {len(MASTER)} entries, 4 sections")


if __name__ == "__main__":
    print("Building TCSW Sponsors Master documents...\n")
    build_xlsx()
    build_pdf()
    print(f"\n✅ Done.")
    print(f"   XLSX: {OUT / 'TCSW_Sponsors_Master.xlsx'}")
    print(f"   PDF:  {OUT / 'TCSW_Sponsors_Master.pdf'}")
