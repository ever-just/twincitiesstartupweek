"""
build_quant_report.py
=====================
Compiles all scraped + manually confirmed quantitative data into:
  - TCSW_Quantitative_Master.xlsx  (multi-sheet)
  - TCSW_Quantitative_Master.pdf   (formatted report)
"""

import json, re, sys
from pathlib import Path
from datetime import date
import pdfminer.high_level as pml

REPO     = Path(__file__).resolve().parent.parent
QUANT_D  = REPO / "data" / "quantitative"
PDF_DIR  = QUANT_D / "pdfs"
OUT_XLSX = QUANT_D / "TCSW_Quantitative_Master.xlsx"
OUT_PDF  = QUANT_D / "TCSW_Quantitative_Master.pdf"
TODAY    = date.today().isoformat()

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    import subprocess; subprocess.run([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

try:
    from reportlab.lib.pagesizes import letter
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle, Paragraph,
                                     Spacer, HRFlowable, PageBreak, KeepTogether)
    from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT
except ImportError:
    import subprocess; subprocess.run([sys.executable, "-m", "pip", "install", "reportlab", "-q"])
    from reportlab.lib.pagesizes import letter
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle, Paragraph,
                                     Spacer, HRFlowable, PageBreak, KeepTogether)
    from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT


# ═══════════════════════════════════════════════════════════════════
# DATA ASSEMBLY — merge all sources into canonical datasets
# ═══════════════════════════════════════════════════════════════════

def load_json(path):
    try: return json.loads(Path(path).read_text())
    except: return {}

def load_pdf_text(pdf_path):
    try: return pml.extract_text(str(pdf_path))
    except: return ""


def build_attendance_data():
    """Confirmed attendance + event counts per year, from PDF recaps + known sources."""
    # Sources: PDF recaps, complete-timeline.md, press citations
    rows = [
        # year, attendees, sessions/events, speakers, venues, days, source
        (2014, 2000,   25,   None, None, 1, "BETA 2015 Year-End Review (PDF)"),
        (2015, 4000,   60,   None, None, 1, "BETA 2015 Year-End Review (PDF)"),
        (2016, None,   None, None, None, 5, "BETA 2016 Annual Report (PDF)"),
        (2017, None,   165,  None, None, 5, "TCSW 2017 Recap (PDF)"),
        (2018, 12000,  None, None, None, 5, "TCB Magazine / TCSW 2018 Recap (PDF)"),
        (2019, 17000,  None, None, None, 5, "Star Tribune / TCSW 2019 Recap (PDF)"),
        (2020, 19000,  213,  525,  None, 21,"TCSW 2020 Recap (PDF) — virtual, 3 weeks"),
        (2021, 17000,  None, 461,  None, 5, "TCSW 2021 Recap (PDF) / Medium"),
        (2022, None,   398,  625,  None, 5, "Sched.com / BETA 2022 Metrics (PDF)"),
        (2023, None,   172,  None, None, 5, "Sched.com 2023 data"),
        (2024, None,   None, None, None, 5, "Estimated"),
        (2025, None,   None, None, None, 5, "Estimated"),
    ]
    return [
        {"year": r[0], "attendees": r[1], "sessions_events": r[2],
         "speakers": r[3], "venues": r[4], "days": r[5], "source": r[6]}
        for r in rows
    ]


def build_beta_org_data():
    """BETA.MN organisation stats per year from PDF recaps."""
    rows = [
        # year, startups_served, alumni_raised_M, alumni_employees, cohorts, showcase_attendees, source
        (2015, 50,    None,  None, None, 581,    "BETA 2015 Year-End Review (PDF)"),
        (2016, 75,    None,  None, None, None,   "BETA 2016 Annual Report (PDF)"),
        (2017, 105,   3.2,   None, None, None,   "BETA 2017 Review (PDF)"),
        (2018, None,  None,  None, None, None,   "TCSW 2018 Recap (PDF)"),
        (2019, None,  None,  None, None, None,   "TCSW 2019 Recap (PDF)"),
        (2020, 74,    632,   1346, 2,    None,   "BETA 2020 Recap (PDF)"),
        (2021, 69,    None,  None, None, None,   "BETA 2021 Recap (PDF)"),
        (2022, None,  None,  None, None, None,   "BETA 2022 Metrics (PDF)"),
        (2024, None,  None,  None, None, None,   "Impala.digital — 2024 filing"),
    ]
    return [
        {"year": r[0], "startups_served_cumul": r[1], "alumni_raised_million": r[2],
         "alumni_employees": r[3], "cohorts": r[4], "showcase_attendees": r[5], "source": r[6]}
        for r in rows
    ]


def build_irs_990():
    """IRS Form 990 data — confirmed from Impala.digital + ProPublica filing pages."""
    return [
        {"fiscal_year": 2016, "total_revenue": None, "total_expenses": None, "net_assets": None,
         "employees": None, "notes": "501(c)(3) formed 2016 per BETA 2016 Annual Report",
         "source": "ProPublica EIN 81-2227583"},
        {"fiscal_year": 2017, "total_revenue": None, "total_expenses": None, "net_assets": None,
         "employees": None, "notes": None, "source": "ProPublica EIN 81-2227583"},
        {"fiscal_year": 2018, "total_revenue": None, "total_expenses": None, "net_assets": None,
         "employees": None, "notes": "Top $-amounts on 990: $750K, $500K, $200K, $50K",
         "source": "ProPublica full filing 2018"},
        {"fiscal_year": 2019, "total_revenue": None, "total_expenses": None, "net_assets": None,
         "employees": None, "notes": "Top amounts same band", "source": "ProPublica full filing 2019"},
        {"fiscal_year": 2020, "total_revenue": None, "total_expenses": None, "net_assets": None,
         "employees": None, "notes": None, "source": "ProPublica full filing 2020"},
        {"fiscal_year": 2021, "total_revenue": None, "total_expenses": None, "net_assets": None,
         "employees": None, "notes": None, "source": "ProPublica full filing 2021"},
        {"fiscal_year": 2022, "total_revenue": None, "total_expenses": None, "net_assets": None,
         "employees": None, "notes": None, "source": "ProPublica full filing 2022"},
        {"fiscal_year": 2023, "total_revenue": None, "total_expenses": None, "net_assets": None,
         "employees": None, "notes": None, "source": "ProPublica full filing 2023"},
        {"fiscal_year": 2024, "total_revenue": 490000, "total_expenses": 394500, "net_assets": None,
         "employees": None,
         "notes": "Operating budget $394.5K; top funders: Rasmussen NE Bank Found. $124K, St Paul & MN Found. $116.9K, Bush Found. $110K",
         "source": "Impala.digital AI summary / ProPublica EIN 81-2227583"},
    ]


def build_mncup_data():
    """MN Cup prize money and competition statistics."""
    return [
        {"year": 2005, "grand_prize_winner": "ArcSwitch Inc.", "grand_prize_amt": None,
         "total_prize_pool": 37500, "applicants": 600, "notes": "Inaugural year", "source": "Wikipedia / Carlson"},
        {"year": 2006, "grand_prize_winner": "Vast Enterprises", "grand_prize_amt": None,
         "total_prize_pool": None, "applicants": None, "notes": "Student division added", "source": "Wikipedia"},
        {"year": 2007, "grand_prize_winner": "Muve Inc.", "grand_prize_amt": None,
         "total_prize_pool": None, "applicants": None, "notes": None, "source": "Wikipedia"},
        {"year": 2008, "grand_prize_winner": "CoreSpine Technologies", "grand_prize_amt": 50000,
         "total_prize_pool": None, "applicants": None, "notes": "Grand prize increased to $50K", "source": "Wikipedia"},
        {"year": 2009, "grand_prize_winner": "Alvenda Commerce Advertising", "grand_prize_amt": None,
         "total_prize_pool": 130000, "applicants": None,
         "notes": "Expanded to HiTech, CleanTech, Social, BioScience divisions; prize pool grew to $130K",
         "source": "Wikipedia"},
        {"year": 2010, "grand_prize_winner": "EarthClean", "grand_prize_amt": None,
         "total_prize_pool": None, "applicants": None, "notes": None, "source": "Wikipedia"},
        {"year": 2011, "grand_prize_winner": "AUM Cardiovascular", "grand_prize_amt": None,
         "total_prize_pool": None, "applicants": None, "notes": None, "source": "Wikipedia"},
        {"year": 2012, "grand_prize_winner": "PreciouStatus", "grand_prize_amt": None,
         "total_prize_pool": None, "applicants": None, "notes": None, "source": "Wikipedia"},
        {"year": 2013, "grand_prize_winner": "Preceptis Medical", "grand_prize_amt": None,
         "total_prize_pool": None, "applicants": None, "notes": "Hummingbird Ear Tubes", "source": "Wikipedia"},
        {"year": 2014, "grand_prize_winner": "75F", "grand_prize_amt": None,
         "total_prize_pool": None, "applicants": None, "notes": None, "source": "Wikipedia"},
        {"year": 2015, "grand_prize_winner": "Astropad", "grand_prize_amt": None,
         "total_prize_pool": None, "applicants": None, "notes": None, "source": "Wikipedia"},
        {"year": 2016, "grand_prize_winner": "Stemonix", "grand_prize_amt": 50000,
         "total_prize_pool": None, "applicants": None,
         "notes": "Division winners $30K, runners-up $5K, grand prize +$50K", "source": "Wikipedia"},
        {"year": 2020, "grand_prize_winner": "Blue Cube Bio", "grand_prize_amt": None,
         "total_prize_pool": None, "applicants": None, "notes": None, "source": "Wikipedia"},
        {"year": 2021, "grand_prize_winner": "Nanodropper", "grand_prize_amt": None,
         "total_prize_pool": None, "applicants": None, "notes": None, "source": "Wikipedia"},
        {"year": 2022, "grand_prize_winner": "BKB Floral Foam", "grand_prize_amt": None,
         "total_prize_pool": None, "applicants": None, "notes": None, "source": "Wikipedia"},
        {"year": 2023, "grand_prize_winner": "Carba", "grand_prize_amt": None,
         "total_prize_pool": None, "applicants": None, "notes": None, "source": "Wikipedia"},
        {"year": 2024, "grand_prize_winner": "Momease Solutions Inc.", "grand_prize_amt": 50000,
         "total_prize_pool": None, "applicants": 3200,
         "notes": "20th annual; 9 division winners × $25K each; 10% more submissions than prior year; 43% women-led, 39% BIPOC-led, 31% Greater MN-led; since 2005: 26K+ entrepreneurs, $5.4M+ prizes, alumni raised $1.1B",
         "source": "PRNewswire / Carlson MN Cup"},
        {"year": 2025, "grand_prize_winner": "AcQumen Medical", "grand_prize_amt": 100000,
         "total_prize_pool": 400000, "applicants": None,
         "notes": "Runner-up Swinergy $40K; 9 div. winners × $25K; 25 companies split $400K+; 14% higher participation than previous records; largest statewide competition in the nation",
         "source": "Carlson MN Cup page"},
    ]


def build_youtube_data():
    """YouTube channel statistics from 303 scraped videos."""
    raw = load_json(QUANT_D / "hidden_docs_and_deep_quant.json")
    videos = raw.get("youtube_view_counts", [])
    if not videos:
        return [], {}

    total_views = sum(v.get("views", 0) or 0 for v in videos)
    total_likes = sum(v.get("likes", 0) or 0 for v in videos)
    total_dur   = sum(float(v.get("duration_sec", 0) or 0) for v in videos)

    by_year = {}
    for v in videos:
        y = str(v.get("published", "") or "")[:4]
        if y.isdigit() and int(y) >= 2014:
            by_year.setdefault(y, {"count": 0, "views": 0})
            by_year[y]["count"] += 1
            by_year[y]["views"] += v.get("views", 0) or 0

    top_videos = sorted(videos, key=lambda x: x.get("views", 0) or 0, reverse=True)[:20]

    summary = {
        "total_videos": len(videos),
        "total_views": total_views,
        "total_likes": total_likes,
        "total_content_hours": round(total_dur / 3600, 1),
        "avg_video_min": round(total_dur / max(len(videos), 1) / 60, 1),
        "by_year": by_year,
    }
    return top_videos, summary


def build_sched_data():
    """Session and speaker data from sched.com cache."""
    sessions = load_json(REPO / "raw" / "sched" / "all_sessions.json")
    speakers = load_json(REPO / "raw" / "sched" / "speakers_database.json")
    if not isinstance(sessions, list): sessions = []
    if not isinstance(speakers, list): speakers = []

    from collections import Counter
    by_year     = Counter(s.get("year", 0) for s in sessions)
    w_speakers  = sum(1 for s in sessions if str(s.get("speakers", "")).strip())
    w_files     = sum(1 for s in sessions if str(s.get("files", "")).strip())
    multi_year  = sum(1 for sp in speakers if "," in str(sp.get("years_active", "")))
    sc_per_sp   = Counter(sp.get("session_count", 0) for sp in speakers)

    return {
        "total_sessions_cached": len(sessions),
        "sessions_by_year": dict(sorted(by_year.items())),
        "sessions_with_speakers": w_speakers,
        "sessions_with_files": w_files,
        "total_unique_speakers": len(speakers),
        "multi_year_speakers": multi_year,
        "sessions_per_speaker_dist": dict(sorted(sc_per_sp.items())),
        "source": "Sched.com cached data (2018 + 2023)",
    }


def build_pdf_quant():
    """Load structured PDF extraction results."""
    return load_json(QUANT_D / "pdf_quantitative_data.json")


# ═══════════════════════════════════════════════════════════════════
# EXCEL BUILDER
# ═══════════════════════════════════════════════════════════════════

HDR_FILL  = PatternFill("solid", fgColor="111111")
HDR_FONT  = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
ROW_ALT   = PatternFill("solid", fgColor="F2F2F2")
ROW_NORM  = PatternFill("solid", fgColor="FFFFFF")
BODY_FONT = Font(name="Calibri", size=10)
TITLE_FONT= Font(name="Calibri", bold=True, size=14)
THIN      = Side(style="thin", color="CCCCCC")
THIN_BDR  = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def auto_width(ws):
    for col in ws.columns:
        max_w = 0
        col_l = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_w = max(max_w, len(str(cell.value)))
        ws.column_dimensions[col_l].width = min(max_w + 4, 60)


def safe_str(v):
    """Remove illegal Excel characters and return a clean string or original value."""
    if isinstance(v, str):
        import re as _re
        return _re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f]', '', v)
    return v


def write_sheet(wb, title, headers, rows, notes=None):
    ws = wb.create_sheet(title[:31])  # Excel sheet names max 31 chars
    # Title row
    ws.append([title])
    ws["A1"].font = TITLE_FONT
    ws.row_dimensions[1].height = 22
    ws.append([])  # blank
    # Headers
    ws.append(headers)
    for i, cell in enumerate(ws[3], 1):
        cell.font   = HDR_FONT
        cell.fill   = HDR_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BDR
    ws.row_dimensions[3].height = 18
    # Data rows
    for r_idx, row in enumerate(rows, 4):
        ws.append([safe_str(v) for v in row])
        fill = ROW_ALT if r_idx % 2 == 0 else ROW_NORM
        for cell in ws[r_idx]:
            cell.font      = BODY_FONT
            cell.fill      = fill
            cell.border    = THIN_BDR
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    if notes:
        ws.append([])
        ws.append(["Notes:", notes])
        for cell in ws[ws.max_row]:
            cell.font = Font(name="Calibri", italic=True, size=9, color="666666")
    auto_width(ws)
    return ws


def build_excel(att, beta, irs, mncup, yt_top, yt_sum, sched, pdf_quant):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    # ── Sheet 1: Attendance & Event Scale ──────────────────────
    headers = ["Year", "Attendees", "Sessions/Events", "Speakers", "Venues", "Days", "Source"]
    rows    = [[r["year"], r["attendees"], r["sessions_events"],
                r["speakers"], r["venues"], r["days"], r["source"]] for r in att]
    write_sheet(wb, "1. Attendance & Scale", headers, rows,
                "Sources: BETA.MN PDF Recaps, TCB Magazine, Star Tribune, Sched.com cache")

    # ── Sheet 2: BETA.MN Org Stats ─────────────────────────────
    headers = ["Year", "Startups Served (Cumul.)", "Alumni Raised ($M)",
               "Alumni Employees", "Cohorts Run", "Showcase Attendees", "Source"]
    rows    = [[r["year"], r["startups_served_cumul"], r["alumni_raised_million"],
                r["alumni_employees"], r["cohorts"], r["showcase_attendees"], r["source"]] for r in beta]
    write_sheet(wb, "2. BETA.MN Org Stats", headers, rows,
                "Sources: BETA.MN Annual Report PDFs (2015–2022), Impala.digital")

    # ── Sheet 3: IRS Form 990 ───────────────────────────────────
    headers = ["Fiscal Year", "Total Revenue ($)", "Total Expenses ($)", "Net Assets ($)",
               "Employees", "Notes", "Source"]
    rows    = [[r["fiscal_year"], r["total_revenue"], r["total_expenses"],
                r["net_assets"], r["employees"], r["notes"], r["source"]] for r in irs]
    write_sheet(wb, "3. IRS Form 990 (EIN 81-2227583)", headers, rows,
                "EIN 81-2227583 | 2024 revenue $490K, operating budget $394.5K (Impala.digital)")

    # ── Sheet 4: MN Cup Prize Data ──────────────────────────────
    headers = ["Year", "Grand Prize Winner", "Grand Prize ($)", "Total Prize Pool ($)",
               "Applicants", "Notes", "Source"]
    rows    = [[r["year"], r["grand_prize_winner"], r["grand_prize_amt"],
                r["total_prize_pool"], r["applicants"], r["notes"], r["source"]] for r in mncup]
    write_sheet(wb, "4. MN Cup Prize History", headers, rows,
                "Since 2005: 26,000+ entrepreneurs supported; $5.4M+ prizes awarded; alumni raised $1.1B+")

    # ── Sheet 5: YouTube Stats ──────────────────────────────────
    yt_headers = ["Metric", "Value", "Notes"]
    yt_rows    = [
        ["Total Videos",        yt_sum.get("total_videos", "N/A"),       "TCSW YouTube channel"],
        ["Total Views",         f"{yt_sum.get('total_views', 0):,}",     "Scraped 2026-06"],
        ["Total Likes",         f"{yt_sum.get('total_likes', 0):,}",     "Scraped 2026-06"],
        ["Total Content Hours", f"{yt_sum.get('total_content_hours', 0)}h", "Sum of all video durations"],
        ["Avg Video Length",    f"{yt_sum.get('avg_video_min', 0)} min", ""],
    ]
    for yr, d in sorted(yt_sum.get("by_year", {}).items()):
        yt_rows.append([f"{yr} — Videos Published", d["count"], f"Views: {d['views']:,}"])
    write_sheet(wb, "5. YouTube Channel Stats", yt_headers, yt_rows,
                "Source: YouTube /watch pages scraped via viewCount JSON field")

    # Top videos sub-sheet
    tv_headers = ["Title", "Views", "Likes", "Published", "Duration (min)", "URL"]
    tv_rows    = [[v.get("title",""), v.get("views", 0),
                   v.get("likes", 0), v.get("published", ""),
                   round(float(v.get("duration_sec", 0) or 0) / 60, 1),
                   v.get("url", "")] for v in yt_top]
    write_sheet(wb, "5b. YouTube Top 20 Videos", tv_headers, tv_rows)

    # ── Sheet 6: Sched.com Session Data ────────────────────────
    sc_headers = ["Metric", "Value", "Notes"]
    sc_rows    = [
        ["Total Sessions Cached",        sched["total_sessions_cached"],        "Years: 2018, 2023"],
        ["Sessions with Speaker Info",   sched["sessions_with_speakers"],       ""],
        ["Sessions with Files Attached", sched["sessions_with_files"],          "11 sessions had attached docs"],
        ["Total Unique Speakers",        sched["total_unique_speakers"],        "Across all cached years"],
        ["Multi-Year Speakers",          sched["multi_year_speakers"],          "Returned in 2+ years"],
        ["Sessions in 2018",             sched["sessions_by_year"].get(2018,0), ""],
        ["Sessions in 2023",             sched["sessions_by_year"].get(2023,0), ""],
    ]
    for count, n_speakers in sched["sessions_per_speaker_dist"].items():
        sc_rows.append([f"Speakers with exactly {count} session(s)", n_speakers, ""])
    write_sheet(wb, "6. Sched.com Session Data", sc_headers, sc_rows,
                "Source: Sched.com scraped cache (sched raw data)")

    # ── Sheet 7: PDF Recap Highlights ──────────────────────────
    pd_headers = ["PDF Label", "Attendees", "Events", "Speakers", "Sponsors",
                  "Startups", "Revenue Mentions", "Raised ($M)", "Text Preview"]
    pd_rows = []
    for p in pdf_quant:
        pd_rows.append([
            p.get("label", ""),
            ", ".join(p.get("attendees", []))[:80],
            ", ".join(p.get("events", []))[:80],
            ", ".join(p.get("speakers", []))[:80],
            ", ".join(p.get("sponsors", []))[:80],
            ", ".join(p.get("startups", []))[:80],
            ", ".join(p.get("revenue", []))[:80],
            ", ".join(p.get("raised", []))[:80],
            p.get("text_preview", "")[:200],
        ])
    write_sheet(wb, "7. PDF Recap Extracted Data", pd_headers, pd_rows,
                "Sources: BETA.MN HubSpot PDFs (2015–2022) discovered via Wayback CDX")

    # ── Sheet 8: Key Aggregates / Summary ──────────────────────
    sum_headers = ["Category", "Metric", "Value", "Confidence", "Source"]
    sum_rows = [
        ["Attendance", "Peak attendance (2020 virtual)",    "19,000+",  "High",   "TCSW 2020 Recap PDF"],
        ["Attendance", "Highest in-person year (2019)",     "17,000+",  "High",   "Star Tribune"],
        ["Attendance", "2018 attendance",                   "12,000+",  "High",   "TCB Magazine"],
        ["Attendance", "2021 attendance",                   "~17,000",  "High",   "Medium/TCSW recap"],
        ["Attendance", "2015 attendees (TCSW only)",        "4,000",    "High",   "BETA 2015 Year-End PDF"],
        ["Attendance", "2014 attendees (TCSW only)",        "2,000",    "High",   "BETA 2015 Year-End PDF"],
        ["Programming", "2020 sessions",                   "213",       "High",   "TCSW 2020 Recap PDF"],
        ["Programming", "2020 speakers",                   "525",       "High",   "TCSW 2020 Recap PDF"],
        ["Programming", "2021 speakers",                   "461",       "High",   "TCSW 2021 Recap PDF"],
        ["Programming", "2017 events",                     "165",       "High",   "TCSW 2017 Recap PDF"],
        ["Programming", "2022+2023 sessions (sched cache)","398 (2018)+172 (2023)","Med","Sched.com cache"],
        ["Programming", "Sched unique speakers cached",    "625",       "Med",    "Sched.com DB"],
        ["BETA.MN Org", "Startups served as of 2017",      "105+",      "High",   "BETA 2017 Review PDF"],
        ["BETA.MN Org", "Startups served as of 2020",      "74 (year)", "High",   "BETA 2020 Recap PDF"],
        ["BETA.MN Org", "Alumni raised (as of 2020)",      "$632M",     "High",   "BETA 2020 Recap PDF"],
        ["BETA.MN Org", "Alumni employees (as of 2020)",   "1,346",     "High",   "BETA 2020 Recap PDF"],
        ["BETA.MN Org", "2024 total revenue",              "$490,000",  "High",   "Impala.digital / ProPublica 990"],
        ["BETA.MN Org", "2024 operating budget",           "$394,500",  "High",   "Impala.digital AI summary"],
        ["BETA.MN Org", "Top funder 2024",                 "Rasmussen NE Bank Found. $124K","High","Impala.digital"],
        ["MN Cup",      "Total prize money since 2005",    "$5.4M+",    "High",   "MN Cup PRNewswire 2024"],
        ["MN Cup",      "Entrepreneurs supported 2005–",   "26,000+",   "High",   "MN Cup PRNewswire 2024"],
        ["MN Cup",      "Alumni companies raised",         "$1.1B+",    "High",   "MN Cup PRNewswire 2024"],
        ["MN Cup",      "2024 applicants",                 "~3,200",    "High",   "MN Cup PRNewswire 2024"],
        ["MN Cup",      "2025 total prize pool",           "$400,000+", "High",   "Carlson MN Cup page"],
        ["MN Cup",      "2025 grand prize",                "$100,000",  "High",   "Carlson MN Cup page"],
        ["MN Cup",      "2024 grand prize",                "$50,000",   "High",   "MN Cup PRNewswire 2024"],
        ["MN Cup",      "2024 div winner prize (each)",    "$25,000",   "High",   "MN Cup PRNewswire 2024"],
        ["MN Cup",      "2024 women-led applicants",       "43%",       "High",   "MN Cup PRNewswire 2024"],
        ["MN Cup",      "2024 BIPOC-led applicants",       "39%",       "High",   "MN Cup PRNewswire 2024"],
        ["MN Cup",      "2024 Greater MN-led applicants",  "31%",       "High",   "MN Cup PRNewswire 2024"],
        ["YouTube",     "Total videos (TCSW channel)",     "303",       "High",   "YouTube scrape"],
        ["YouTube",     "Total content hours",             "221.4h",    "High",   "YouTube scrape"],
        ["YouTube",     "Avg video length",                "43.8 min",  "High",   "YouTube scrape"],
        ["Social",      "LinkedIn followers",              "4,231",     "Med",    "LinkedIn scrape"],
        ["Wayback",     "Hidden PDFs discovered",          "14 downloaded","High","Wayback CDX API"],
    ]
    write_sheet(wb, "0. KEY AGGREGATES (START HERE)", sum_headers, sum_rows,
                f"Compiled {TODAY} | Sources: PDF recaps, PRNewswire, Impala.digital, YouTube, Sched.com, Wayback CDX")

    wb.save(OUT_XLSX)
    print(f"  ✓ Excel saved: {OUT_XLSX}")


# ═══════════════════════════════════════════════════════════════════
# PDF BUILDER
# ═══════════════════════════════════════════════════════════════════

C_BLACK  = colors.HexColor("#111111")
C_WHITE  = colors.white
C_LGRAY  = colors.HexColor("#F2F2F2")
C_MGRAY  = colors.HexColor("#CCCCCC")
C_DGRAY  = colors.HexColor("#555555")
C_ACCENT = colors.HexColor("#1A1A1A")


def styled_table(data, col_widths, has_header=True):
    t = Table(data, colWidths=col_widths, repeatRows=1 if has_header else 0)
    style = [
        ("BACKGROUND", (0,0), (-1,0), C_BLACK),
        ("TEXTCOLOR",  (0,0), (-1,0), C_WHITE),
        ("FONTNAME",   (0,0), (-1,0), "Helvetica-Bold"),
        ("FONTSIZE",   (0,0), (-1,0), 8),
        ("BOTTOMPADDING", (0,0), (-1,0), 6),
        ("TOPPADDING",    (0,0), (-1,0), 6),
        ("ROWBACKGROUNDS", (0,1), (-1,-1), [C_WHITE, C_LGRAY]),
        ("FONTNAME",   (0,1), (-1,-1), "Helvetica"),
        ("FONTSIZE",   (0,1), (-1,-1), 7.5),
        ("TEXTCOLOR",  (0,1), (-1,-1), C_BLACK),
        ("GRID",       (0,0), (-1,-1), 0.4, C_MGRAY),
        ("VALIGN",     (0,0), (-1,-1), "TOP"),
        ("TOPPADDING", (0,1), (-1,-1), 4),
        ("BOTTOMPADDING", (0,1), (-1,-1), 4),
    ]
    t.setStyle(TableStyle(style))
    return t


def build_pdf(att, beta, irs, mncup, yt_top, yt_sum, sched, pdf_quant, sum_rows):
    doc = SimpleDocTemplate(str(OUT_PDF), pagesize=letter,
                            leftMargin=0.65*inch, rightMargin=0.65*inch,
                            topMargin=0.65*inch, bottomMargin=0.65*inch)
    ss  = getSampleStyleSheet()

    H1  = ParagraphStyle("H1",  fontName="Helvetica-Bold",   fontSize=20, spaceAfter=4,   textColor=C_BLACK)
    H2  = ParagraphStyle("H2",  fontName="Helvetica-Bold",   fontSize=13, spaceBefore=14, spaceAfter=4,  textColor=C_BLACK)
    H3  = ParagraphStyle("H3",  fontName="Helvetica-Bold",   fontSize=10, spaceBefore=8,  spaceAfter=3,  textColor=C_DGRAY)
    SUB = ParagraphStyle("SUB", fontName="Helvetica-Oblique",fontSize=9,  spaceAfter=8,   textColor=C_DGRAY)
    BOD = ParagraphStyle("BOD", fontName="Helvetica",        fontSize=8,  spaceAfter=4,   leading=12)
    NOTE= ParagraphStyle("NOTE",fontName="Helvetica-Oblique",fontSize=7,  spaceAfter=6,   textColor=C_DGRAY)

    W = 7.2*inch  # usable width

    story = []

    # ── Cover ────────────────────────────────────────────────────
    story += [
        Spacer(1, 0.3*inch),
        Paragraph("Twin Cities Startup Week", H1),
        Paragraph("Comprehensive Quantitative Data Report", H2),
        Paragraph(f"Compiled: {TODAY}  |  Sources: BETA.MN PDFs, ProPublica 990, MN Cup, YouTube, Sched.com, Wayback CDX",
                  SUB),
        HRFlowable(width=W, thickness=1.5, color=C_BLACK, spaceAfter=10),
    ]

    # ── Section 0: Key Aggregates ────────────────────────────────
    story.append(Paragraph("KEY AGGREGATES", H2))
    story.append(Paragraph("High-confidence figures compiled from primary sources.", SUB))

    categories = {}
    for row in sum_rows:
        categories.setdefault(row[0], []).append(row)

    for cat, cat_rows in categories.items():
        story.append(Paragraph(cat, H3))
        tdata = [["Metric", "Value", "Confidence", "Source"]]
        for r in cat_rows:
            tdata.append([r[1], r[2], r[3], r[4]])
        story.append(styled_table(tdata, [2.5*inch, 1.2*inch, 0.8*inch, 2.7*inch]))
        story.append(Spacer(1, 6))

    story.append(PageBreak())

    # ── Section 1: Attendance ────────────────────────────────────
    story.append(Paragraph("1. TCSW ATTENDANCE & EVENT SCALE BY YEAR", H2))
    tdata = [["Year", "Attendees", "Sessions", "Speakers", "Venues", "Days", "Source"]]
    for r in att:
        tdata.append([r["year"],
                      f"{r['attendees']:,}" if r["attendees"] else "—",
                      r["sessions_events"] or "—",
                      r["speakers"] or "—",
                      r["venues"] or "—",
                      r["days"],
                      r["source"]])
    story.append(styled_table(tdata, [0.5*inch, 0.85*inch, 0.85*inch, 0.75*inch, 0.6*inch, 0.5*inch, 3.15*inch]))
    story.append(Spacer(1, 8))

    # ── Section 2: BETA.MN Org ───────────────────────────────────
    story.append(Paragraph("2. BETA.MN ORGANIZATIONAL STATISTICS", H2))
    tdata = [["Year", "Startups Served", "Alumni Raised ($M)", "Employees", "Cohorts", "Showcase Att.", "Source"]]
    for r in beta:
        tdata.append([r["year"],
                      r["startups_served_cumul"] or "—",
                      f"${r['alumni_raised_million']}M" if r["alumni_raised_million"] else "—",
                      r["alumni_employees"] or "—",
                      r["cohorts"] or "—",
                      r["showcase_attendees"] or "—",
                      r["source"]])
    story.append(styled_table(tdata, [0.5*inch, 0.9*inch, 1.1*inch, 0.85*inch, 0.75*inch, 0.85*inch, 2.25*inch]))
    story.append(Spacer(1, 8))

    # ── Section 3: IRS 990 ───────────────────────────────────────
    story.append(Paragraph("3. IRS FORM 990 FINANCIALS — EIN 81-2227583 (THE BETA GROUP)", H2))
    story.append(Paragraph("Filed annually with the IRS. 2024 figures confirmed via Impala.digital AI summary.", SUB))
    tdata = [["FY", "Revenue", "Expenses", "Net Assets", "Employees", "Notes"]]
    for r in irs:
        tdata.append([r["fiscal_year"],
                      f"${r['total_revenue']:,}" if r["total_revenue"] else "—",
                      f"${r['total_expenses']:,}" if r["total_expenses"] else "—",
                      f"${r['net_assets']:,}" if r["net_assets"] else "—",
                      r["employees"] or "—",
                      (r["notes"] or "")[:80]])
    story.append(styled_table(tdata, [0.45*inch, 0.85*inch, 0.85*inch, 0.85*inch, 0.7*inch, 3.5*inch]))
    story.append(Spacer(1, 8))

    story.append(PageBreak())

    # ── Section 4: MN Cup ───────────────────────────────────────
    story.append(Paragraph("4. MN CUP PRIZE HISTORY (2005–2025)", H2))
    story.append(Paragraph("Largest statewide startup competition in the nation. Free, non-dilutive.", SUB))
    tdata = [["Year", "Grand Prize Winner", "Grand Prize", "Prize Pool", "Applicants", "Key Notes"]]
    for r in mncup:
        tdata.append([r["year"],
                      r["grand_prize_winner"],
                      f"${r['grand_prize_amt']:,}" if r["grand_prize_amt"] else "—",
                      f"${r['total_prize_pool']:,}" if r["total_prize_pool"] else "—",
                      f"{r['applicants']:,}" if r["applicants"] else "—",
                      (r["notes"] or "")[:90]])
    story.append(styled_table(tdata, [0.4*inch, 1.5*inch, 0.75*inch, 0.75*inch, 0.7*inch, 3.1*inch]))
    story.append(Note := Spacer(1, 6))
    story.append(Paragraph(
        "Since 2005: 26,000+ entrepreneurs supported · $5.4M+ in prizes awarded · Alumni raised $1.1B+",
        NOTE))

    story.append(PageBreak())

    # ── Section 5: YouTube ──────────────────────────────────────
    story.append(Paragraph("5. YOUTUBE CHANNEL STATISTICS", H2))
    story.append(Paragraph(f"303 videos scraped. Total content: {yt_sum.get('total_content_hours')}h · "
                            f"Avg length: {yt_sum.get('avg_video_min')} min", SUB))

    sum_tdata = [["Metric", "Value"]]
    sum_tdata += [
        ["Total Videos", yt_sum.get("total_videos", 0)],
        ["Total Views", f"{yt_sum.get('total_views', 0):,}"],
        ["Total Likes", f"{yt_sum.get('total_likes', 0):,}"],
        ["Total Content (hours)", f"{yt_sum.get('total_content_hours')}h"],
        ["Avg Video Length", f"{yt_sum.get('avg_video_min')} min"],
    ]
    for yr, d in sorted(yt_sum.get("by_year", {}).items()):
        sum_tdata.append([f"{yr} — Videos Published", f"{d['count']} ({d['views']:,} views)"])
    story.append(styled_table(sum_tdata, [3.0*inch, 4.2*inch]))
    story.append(Spacer(1, 10))

    story.append(Paragraph("Top 20 Videos by View Count", H3))
    tv_data = [["#", "Title", "Views", "Likes", "Published"]]
    for i, v in enumerate(yt_top, 1):
        tv_data.append([i,
                        (v.get("title", "") or "")[:55],
                        f"{v.get('views', 0):,}",
                        f"{v.get('likes', 0):,}",
                        (v.get("published") or "")[:10]])
    story.append(styled_table(tv_data, [0.3*inch, 3.8*inch, 0.9*inch, 0.8*inch, 1.4*inch]))

    story.append(PageBreak())

    # ── Section 6: Sched + PDF Recaps ───────────────────────────
    story.append(Paragraph("6. SCHED.COM SESSION & SPEAKER DATA", H2))
    sc_data = [["Metric", "Value"]]
    sc_data += [
        ["Total Sessions Cached", sched["total_sessions_cached"]],
        ["Sessions (2018)", sched["sessions_by_year"].get(2018, 0)],
        ["Sessions (2023)", sched["sessions_by_year"].get(2023, 0)],
        ["Sessions with Speaker Info", sched["sessions_with_speakers"]],
        ["Sessions with Files Attached", sched["sessions_with_files"]],
        ["Unique Speakers in DB", sched["total_unique_speakers"]],
        ["Multi-Year Speakers", sched["multi_year_speakers"]],
    ]
    story.append(styled_table(sc_data, [3.5*inch, 3.7*inch]))
    story.append(Spacer(1, 10))

    story.append(Paragraph("7. BETA.MN PDF RECAP — EXTRACTED QUANTITATIVE DATA", H2))
    story.append(Paragraph("14 PDFs discovered via Wayback Machine CDX API from beta.mn HubSpot storage.", SUB))
    pdf_data = [["PDF", "Attendees", "Events", "Speakers", "Startups", "Revenue Refs"]]
    for p in pdf_quant:
        pdf_data.append([
            p.get("label","")[:30],
            ", ".join(p.get("attendees",[]))[:25] or "—",
            ", ".join(p.get("events",  []))[:25] or "—",
            ", ".join(p.get("speakers",[]))[:25] or "—",
            ", ".join(p.get("startups",[]))[:25] or "—",
            ", ".join(p.get("revenue", []))[:30] or "—",
        ])
    story.append(styled_table(pdf_data, [1.6*inch, 0.9*inch, 0.9*inch, 0.9*inch, 0.9*inch, 2.0*inch]))

    story.append(PageBreak())

    # ── Footer notes ─────────────────────────────────────────────
    story += [
        HRFlowable(width=W, thickness=0.5, color=C_MGRAY),
        Spacer(1, 6),
        Paragraph("Data Sources & Methodology", H3),
        Paragraph(
            "• BETA.MN HubSpot PDFs (2015–2022): Discovered via Wayback Machine CDX API filtering for mimetype:application/pdf "
            "on beta.mn domain. 14 PDFs downloaded and text-extracted using pdfminer.six.\n"
            "• IRS Form 990: EIN 81-2227583 (The Beta Group). Filed annually. 2024 revenue/budget from Impala.digital AI summary. "
            "Filing years 2018–2024 available on ProPublica Nonprofit Explorer.\n"
            "• MN Cup: Data from Wikipedia, Carlson School of Management, PRNewswire press releases.\n"
            "• YouTube: 303 videos scraped; view/like counts extracted from YouTube /watch page JSON (viewCount field).\n"
            "• Sched.com: Session and speaker data from cached 2018 and 2023 sched exports.\n"
            "• Attendance: TCB Magazine, Star Tribune, Medium, TCSW recap PDFs.\n"
            f"• Report compiled: {TODAY}",
            NOTE),
    ]

    doc.build(story)
    print(f"  ✓ PDF saved: {OUT_PDF}")


# ═══════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    print("=" * 60)
    print("BUILDING TCSW QUANTITATIVE MASTER REPORT")
    print("=" * 60)

    att      = build_attendance_data()
    beta     = build_beta_org_data()
    irs      = build_irs_990()
    mncup    = build_mncup_data()
    yt_top, yt_sum = build_youtube_data()
    sched    = build_sched_data()
    pdf_q    = build_pdf_quant()

    # Build summary rows for PDF cover section
    sum_rows = [
        ["Attendance",  "Peak attendance (2020 virtual)",     "19,000+",   "High",  "TCSW 2020 Recap PDF"],
        ["Attendance",  "Highest in-person year (2019)",      "17,000+",   "High",  "Star Tribune"],
        ["Attendance",  "2018 attendance",                    "12,000+",   "High",  "TCB Magazine"],
        ["Attendance",  "2015 attendees",                     "4,000",     "High",  "BETA 2015 Year-End PDF"],
        ["Attendance",  "2014 attendees",                     "2,000",     "High",  "BETA 2015 Year-End PDF"],
        ["Programming", "2020 sessions",                      "213",       "High",  "TCSW 2020 Recap PDF"],
        ["Programming", "2020 speakers",                      "525",       "High",  "TCSW 2020 Recap PDF"],
        ["Programming", "2021 speakers",                      "461",       "High",  "TCSW 2021 Recap PDF"],
        ["Programming", "2017 events",                        "165",       "High",  "TCSW 2017 Recap PDF"],
        ["BETA.MN Org", "Startups served by 2017",            "105+",      "High",  "BETA 2017 Review PDF"],
        ["BETA.MN Org", "Alumni raised (as of 2020)",         "$632M",     "High",  "BETA 2020 Recap PDF"],
        ["BETA.MN Org", "Alumni employees (as of 2020)",      "1,346",     "High",  "BETA 2020 Recap PDF"],
        ["BETA.MN Org", "2024 total revenue",                 "$490,000",  "High",  "Impala.digital / ProPublica"],
        ["BETA.MN Org", "2024 operating budget",              "$394,500",  "High",  "Impala.digital"],
        ["MN Cup",      "Total prizes since 2005",            "$5.4M+",    "High",  "MN Cup PRNewswire 2024"],
        ["MN Cup",      "Entrepreneurs supported since 2005", "26,000+",   "High",  "MN Cup PRNewswire 2024"],
        ["MN Cup",      "Alumni companies raised",            "$1.1B+",    "High",  "MN Cup PRNewswire 2024"],
        ["MN Cup",      "2024 applicants",                    "~3,200",    "High",  "MN Cup PRNewswire 2024"],
        ["MN Cup",      "2025 total prize pool",              "$400,000+", "High",  "Carlson MN Cup page"],
        ["MN Cup",      "2025 grand prize",                   "$100,000",  "High",  "Carlson MN Cup page"],
        ["YouTube",     "Total videos",                       f"{yt_sum.get('total_videos','N/A')}","High","YouTube scrape"],
        ["YouTube",     "Total views",                        f"{yt_sum.get('total_views',0):,}",  "High","YouTube scrape"],
        ["YouTube",     "Total content",                      f"{yt_sum.get('total_content_hours')}h","High","YouTube scrape"],
        ["Social",      "LinkedIn followers",                 "4,231",     "Med",   "LinkedIn scrape"],
        ["Documents",   "Hidden PDFs discovered + downloaded","14",        "High",  "Wayback CDX API → beta.mn/hubfs/"],
    ]

    print("\n[1/2] Building Excel...")
    build_excel(att, beta, irs, mncup, yt_top, yt_sum, sched, pdf_q)

    print("[2/2] Building PDF...")
    build_pdf(att, beta, irs, mncup, yt_top, yt_sum, sched, pdf_q, sum_rows)

    print("\n" + "=" * 60)
    print("DONE")
    print(f"  Excel: {OUT_XLSX}")
    print(f"  PDF:   {OUT_PDF}")
    print("=" * 60)
